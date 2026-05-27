"""Comprehensive 29-field QC scan across all 112,865 unit rows in canary 1ef1060.

For each field, identifies:
- Missing / null / empty rate
- Invalid values (non-numeric where expected, out-of-range)
- Anomalous patterns (duplicates, collisions, divergences)
- Per-tier breakdown of the worst offenders

Output:
  investigations/2026-05-25-unit-debug/FULL_QC_REPORT.md
"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

XLSX = "/Users/ankur/Downloads/canary_1ef1060_2026-05-25_post_phase16_v2.xlsx"
OUT = Path(__file__).parent / "FULL_QC_REPORT.md"


def main() -> None:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Units"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}

    total = 0
    # Per-field metrics
    null_count = Counter()
    invalid_value = defaultdict(Counter)  # field -> {issue: count}
    by_tier = defaultdict(lambda: defaultdict(int))

    # Cross-field
    unit_id_per_prop: dict[int, list[str]] = defaultdict(list)
    fp_id_to_name: dict[str, set] = defaultdict(set)
    name_to_fp_id: dict[str, set] = defaultdict(set)
    avail_raw_vs_postfix_divergence = 0
    huge_cost_per_sqft = []
    units_per_prop: dict[int, int] = defaultdict(int)
    dup_unit_within_prop = defaultdict(int)

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        rec = {h: row[i] for i, h in enumerate(hdr)}
        tier = rec["extraction_tier"] or "?"
        pid = rec["apartment_id"]
        units_per_prop[pid] += 1

        # ─── Per-field null check ──────────────────────────────────────
        for f in hdr:
            v = rec[f]
            is_null = v is None or v == "" or (isinstance(v, str) and not v.strip())
            if is_null:
                null_count[f] += 1

        # ─── unit_id duplicates within property ────────────────────────
        uid = rec.get("unit_id")
        if uid:
            if uid in unit_id_per_prop[pid]:
                dup_unit_within_prop[pid] += 1
            else:
                unit_id_per_prop[pid].append(uid)

        # ─── floor_plan_id ↔ name consistency ──────────────────────────
        fp_id = rec.get("floor_plan_id")
        fp_name = rec.get("floor_plan_name")
        if fp_id and fp_name:
            fp_id_to_name[str(fp_id)].add(str(fp_name))
            name_to_fp_id[str(fp_name)].add(str(fp_id))

        # ─── beds validity ─────────────────────────────────────────────
        beds = rec.get("beds")
        if beds is not None:
            try:
                b = float(beds)
                if b < 0 or b > 10:
                    invalid_value["beds"]["out_of_range"] += 1
                    by_tier[tier]["beds_out_of_range"] += 1
            except (TypeError, ValueError):
                invalid_value["beds"]["non_numeric"] += 1
                by_tier[tier]["beds_non_numeric"] += 1

        # ─── baths validity ────────────────────────────────────────────
        baths = rec.get("baths")
        if baths is not None and baths != "":
            try:
                bath = float(baths)
                if bath < 0 or bath > 10:
                    invalid_value["baths"]["out_of_range"] += 1
                # Common bath values are integers or .5; flag others
                frac = bath - int(bath)
                if frac not in (0.0, 0.5, 0.25, 0.75):
                    invalid_value["baths"]["non_half"] += 1
            except (TypeError, ValueError):
                invalid_value["baths"]["non_numeric"] += 1

        # ─── area_sqft validity ────────────────────────────────────────
        sqft = rec.get("area_sqft")
        if sqft is not None:
            try:
                s = float(sqft)
                if s == -1:
                    invalid_value["area_sqft"]["sentinel_-1"] += 1
                elif s == 0:
                    invalid_value["area_sqft"]["zero"] += 1
                elif s < 100:
                    invalid_value["area_sqft"]["implausibly_small"] += 1
                elif s > 10000:
                    invalid_value["area_sqft"]["implausibly_large"] += 1
            except (TypeError, ValueError):
                invalid_value["area_sqft"]["non_numeric"] += 1

        # ─── rent_low / rent_high ──────────────────────────────────────
        rl, rh = rec.get("rent_low"), rec.get("rent_high")
        if rl is not None:
            try:
                rl_f = float(rl)
                if rl_f == 0:
                    invalid_value["rent_low"]["zero"] += 1
                elif rl_f < 100:
                    invalid_value["rent_low"]["implausibly_small"] += 1
                elif rl_f > 50000:
                    invalid_value["rent_low"]["implausibly_large"] += 1
            except (TypeError, ValueError):
                invalid_value["rent_low"]["non_numeric"] += 1
        if rl and rh:
            try:
                if float(rh) < float(rl):
                    invalid_value["rent_high"]["less_than_low"] += 1
            except (TypeError, ValueError):
                pass
        # Rent spread check
        if rl and rh:
            try:
                spread = float(rh) - float(rl)
                if spread > 5000:
                    invalid_value["rent_spread"]["over_5k"] += 1
            except (TypeError, ValueError):
                pass

        # ─── cost-per-sqft sanity (after we have both) ──────────────────
        try:
            if rl and sqft and sqft != -1 and float(sqft) > 0:
                cps = float(rl) / float(sqft)
                if cps > 20:  # > $20/sqft/mo is luxury
                    invalid_value["rent_per_sqft"]["over_20_usd"] += 1
                if cps < 0.3:  # < $0.30/sqft/mo is implausibly cheap
                    invalid_value["rent_per_sqft"]["under_0.3_usd"] += 1
        except (TypeError, ValueError):
            pass

        # ─── availability_date raw vs post_fix ──────────────────────────
        ar, ap = rec.get("available_date_raw"), rec.get("available_date_post_fix")
        if ar and ap and str(ar).strip() != str(ap).strip():
            avail_raw_vs_postfix_divergence += 1

        # ─── concession fields ─────────────────────────────────────────
        for cf in ("concession_text", "concession_text_clean", "concession_value"):
            if not rec.get(cf):
                null_count[f"concession::{cf}_empty"] += 1

        # ─── offer fields ──────────────────────────────────────────────
        for of in ("offer_banner", "offer_type", "offer_target", "offer_value"):
            if not rec.get(of):
                null_count[f"offer::{of}_empty"] += 1

    # ─── Plan-id ↔ name collision analysis ────────────────────────────
    fp_id_with_multiple_names = sum(1 for v in fp_id_to_name.values() if len(v) > 1)
    name_with_multiple_fp_ids = sum(1 for v in name_to_fp_id.values() if len(v) > 1)

    # ─── Duplicate unit_id per property ───────────────────────────────
    props_with_dups = len([p for p, c in dup_unit_within_prop.items() if c > 0])
    total_dup_unit_ids = sum(dup_unit_within_prop.values())

    # ─── Write report ────────────────────────────────────────────────
    lines = [f"# Full QC Report — canary 1ef1060 ({total} units)", ""]

    lines.append("## Field null/empty rate (all 29 fields)")
    lines.append("| Field | Null/empty | % |")
    lines.append("|---|---:|---:|")
    for f in hdr:
        c = null_count.get(f, 0)
        if c > 0:
            lines.append(f"| `{f}` | {c} | {100*c/total:.1f}% |")
    lines.append("")

    lines.append("## Per-field invalid values")
    for field, issues in sorted(invalid_value.items()):
        lines.append(f"### `{field}`")
        for issue, count in issues.most_common():
            lines.append(f"- {issue}: **{count}** ({100*count/total:.2f}%)")
        lines.append("")

    lines.append("## Cross-field anomalies")
    lines.append(f"- **floor_plan_id with multiple distinct names**: {fp_id_with_multiple_names}")
    lines.append(f"- **floor_plan_name with multiple distinct ids**: {name_with_multiple_fp_ids}")
    lines.append(f"- **available_date_raw vs post_fix divergence**: {avail_raw_vs_postfix_divergence} units (Phase-16 rewrite count)")
    lines.append(f"- **duplicate unit_id within property**: {total_dup_unit_ids} dup rows across {props_with_dups} props")
    lines.append("")

    lines.append("## Concession-field empty rates (5 fields)")
    for cf in ("concession_text", "concession_text_clean", "concession_value"):
        k = f"concession::{cf}_empty"
        c = null_count.get(k, 0)
        lines.append(f"- `{cf}`: {c} empty ({100*c/total:.0f}%)")
    lines.append("")

    lines.append("## Offer-field empty rates (5 fields)")
    for of in ("offer_banner", "offer_type", "offer_target", "offer_value"):
        k = f"offer::{of}_empty"
        c = null_count.get(k, 0)
        lines.append(f"- `{of}`: {c} empty ({100*c/total:.0f}%)")

    OUT.write_text("\n".join(lines))
    print(f"WROTE {OUT}")
    print(f"\n--- TOP 12 NULL FIELDS ---")
    for f, c in null_count.most_common(12):
        print(f"  {c:>7}  ({100*c/total:>5.1f}%)  {f}")
    print(f"\n--- ALL INVALID-VALUE BUCKETS ---")
    for field, issues in sorted(invalid_value.items()):
        for issue, count in issues.most_common():
            print(f"  {count:>7}  ({100*count/total:>5.2f}%)  {field}::{issue}")
    print(f"\n--- CROSS-FIELD ---")
    print(f"  fp_id with multiple names: {fp_id_with_multiple_names}")
    print(f"  fp_name with multiple ids: {name_with_multiple_fp_ids}")
    print(f"  avail_date raw≠postfix:    {avail_raw_vs_postfix_divergence}")
    print(f"  unit_id duplicates:        {total_dup_unit_ids} rows across {props_with_dups} props")


if __name__ == "__main__":
    main()
