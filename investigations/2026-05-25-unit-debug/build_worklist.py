"""Build stratified 40-prop worklist from canary 1ef1060 n_full_post_fix=0 cohort.

Output: artifacts/probe/n_full_zero_worklist.jsonl (one prop per line).
Then a second 40-prop worklist for the low-strict cohort (n_units>0 but
strict_pass_pct<80%) goes to artifacts/probe/low_strict_worklist.jsonl.

Stratification ignores tiers already owned by active chips:
- TIER_1_DOM_ENTRATA_PP_SSR_PLAN_LEVEL (chip #98)
- TIER_1_API_RENTMANAGER (already shipped)
- TIER_1_API_REPLI360_PLAN_LEVEL (already shipped)
"""
from __future__ import annotations

import json
import random
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

XLSX = "/Users/ankur/Downloads/canary_1ef1060_2026-05-25_post_phase16_v2.xlsx"
OUT = Path(__file__).parent / "artifacts" / "probe"

# Tiers owned by active chips/shipped - skip
SKIP_TIERS = {
    "TIER_1_DOM_ENTRATA_PP_SSR_PLAN_LEVEL",  # chip #98
    "TIER_1_API_RENTMANAGER",                 # shipped
    "TIER_1_API_REPLI360_PLAN_LEVEL",         # shipped
}

# Wave 2 (2026-05-25 post-merge): triple the per-tier quota to cover
# 120 fresh props from the remaining ~894 n_full=0 props. Same skip
# logic — tiers owned by in-flight chips are excluded.
# Target sample counts per tier (n_full=0 cohort, 40 total)
N_FULL_ZERO_QUOTA = {
    "TIER_1_API": 5,
    "TIER_3_DOM": 5,
    "TIER_1_API_RENTCAFE_SHAPE_REJECTED": 4,
    "TIER_1_KNOCK_API": 4,
    "generic:no_body_short_circuit": 3,
    "TIER_1_DOM_GENERIC_PLAN_TEXT_PLAN_LEVEL": 3,
    "?": 3,
    "TIER_MERGED_CROSS_PAGE": 2,
    "generic:sgcaptcha_wall": 2,
    "TIER_1_DOM_REALPAGE_CWS": 2,
    "SYNDICATION_ONLY_WIX": 2,
    "TIER_1_DOM_MARKETAPTS": 2,
    "TIER_1_API_APPFOLIO": 2,
    "NOT_ENCORESKYLINE_TEMPLATE": 1,
}


def main() -> None:
    random.seed(42)  # deterministic sampling for reproducibility
    OUT.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Properties"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}

    n_full_zero_by_tier: dict[str, list[dict]] = defaultdict(list)
    partials_by_tier: dict[str, list[dict]] = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: row[i] for i, h in enumerate(hdr)}
        tier = rec.get("extraction_tier") or "?"
        if tier in SKIP_TIERS:
            continue
        n_units = rec.get("n_units") or 0
        n_full = rec.get("n_full_post_fix") or 0

        if n_full == 0:
            n_full_zero_by_tier[tier].append(rec)
        elif n_units > 0 and (n_full / n_units) < 0.80:
            # Low-strict: emit units but <80% are full-row
            partials_by_tier[tier].append(rec)

    # n_full=0 stratified sample
    n_full_zero_sample = []
    for tier, quota in N_FULL_ZERO_QUOTA.items():
        pool = n_full_zero_by_tier.get(tier, [])
        if not pool:
            print(f"⚠ no props in n_full=0 cohort for tier {tier}")
            continue
        picked = random.sample(pool, min(quota, len(pool)))
        for r in picked:
            r["_probe_cohort"] = "n_full_zero"
            r["_probe_tier"] = tier
            r["_pool_size"] = len(pool)
            n_full_zero_sample.append(r)
        print(f"  {len(picked):>2}/{quota} from tier {tier} (pool={len(pool)})")

    with (OUT / "n_full_zero_worklist.jsonl").open("w") as f:
        for r in n_full_zero_sample:
            f.write(json.dumps(r, default=str) + "\n")
    print(f"\nWROTE n_full_zero_worklist.jsonl: {len(n_full_zero_sample)} props\n")

    # Low-strict cohort - same stratification approach, pick top tiers
    LOW_STRICT_QUOTA = {
        # Top low-strict tiers (will be filled below from observed pools)
    }
    # Auto-pick from largest pools that aren't already covered
    sorted_pools = sorted(partials_by_tier.items(), key=lambda kv: -len(kv[1]))
    print("--- Low-strict pools (top 20) ---")
    for tier, pool in sorted_pools[:20]:
        print(f"  {len(pool):>4}  {tier}")

    # Quota: cover top tiers, biggest first; cap total ~40
    total = 0
    auto_quota = {}
    for tier, pool in sorted_pools:
        if total >= 40:
            break
        q = min(4, max(2, len(pool) // 10), len(pool))
        if total + q > 40:
            q = 40 - total
        auto_quota[tier] = q
        total += q

    low_strict_sample = []
    print("\n--- Low-strict sample picks ---")
    for tier, quota in auto_quota.items():
        pool = partials_by_tier[tier]
        picked = random.sample(pool, min(quota, len(pool)))
        for r in picked:
            r["_probe_cohort"] = "low_strict"
            r["_probe_tier"] = tier
            r["_pool_size"] = len(pool)
            low_strict_sample.append(r)
        print(f"  {len(picked):>2}/{quota} from tier {tier} (pool={len(pool)})")

    with (OUT / "low_strict_worklist.jsonl").open("w") as f:
        for r in low_strict_sample:
            f.write(json.dumps(r, default=str) + "\n")
    print(f"\nWROTE low_strict_worklist.jsonl: {len(low_strict_sample)} props")


if __name__ == "__main__":
    main()
