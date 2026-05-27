"""Wave 2 worklist — 120 fresh n_full=0 props.

Skips the 40 props already probed in wave 1. Triples per-tier quota
(plus a few NEW tier signatures uncovered in wave 1 like 365res,
SightMap_IFRAME etc.). Now-shipped tiers from chips are excluded.
"""
from __future__ import annotations

import json
import random
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

XLSX = "/Users/ankur/Downloads/canary_1ef1060_2026-05-25_post_phase16_v2.xlsx"
OUT = Path(__file__).parent / "artifacts" / "probe"
WAVE1_RESULTS = OUT / "n_full_zero_results.jsonl"

# Now-shipped (post-merge 2026-05-25) — exclude from wave 2 sample
SKIP_TIERS = {
    # Wave 1 (or earlier) — already shipped via chips
    "TIER_1_DOM_ENTRATA_PP_SSR_PLAN_LEVEL",  # chip #98 merged
    "TIER_1_API_RENTMANAGER",                 # shipped earlier
    "TIER_1_API_REPLI360_PLAN_LEVEL",         # shipped earlier
}

# Triple wave-1 quotas, add NEW tier signatures discovered in wave 1
N_FULL_ZERO_QUOTA = {
    # Big un-shipped pools (3× wave 1)
    "TIER_1_API": 15,
    "TIER_3_DOM": 15,
    "TIER_1_API_RENTCAFE_SHAPE_REJECTED": 12,
    "TIER_1_API_RENTCAFE_SECURECAFE": 12,  # NEW IN WAVE 2 (55-prop pool, wasn't sampled in wave 1)
    "TIER_1_KNOCK_API": 12,
    "generic:no_body_short_circuit": 9,
    "TIER_1_DOM_GENERIC_PLAN_TEXT_PLAN_LEVEL": 9,
    "?": 9,
    "TIER_MERGED_CROSS_PAGE": 6,
    "generic:sgcaptcha_wall": 6,
    "TIER_1_API_ENTRATA_SHAPE_REJECTED": 6,  # NEW IN WAVE 2 (19-prop pool)
    "TIER_1_API_ENTRATA_EMPTY": 6,            # NEW IN WAVE 2 (18-prop pool)
    "TIER_1_DOM_REALPAGE_CWS": 6,
    "SYNDICATION_ONLY_WIX": 6,
    "TIER_1_DOM_MARKETAPTS": 6,
    "TIER_1_API_APPFOLIO": 6,
    "SYNDICATION_ONLY_SQUARESPACE": 4,        # NEW IN WAVE 2 (12-prop pool)
    "TIER_1_API_ENTRATA_NO_RESPONSE": 4,      # NEW IN WAVE 2 (12-prop pool)
    "TIER_1_API_ONESITE_NO_RESPONSE": 4,      # NEW IN WAVE 2 (10-prop pool)
    "TIER_1_API_SIGHTMAP_IFRAME": 4,          # NEW IN WAVE 2 (9-prop pool)
    "NOT_ENCORESKYLINE_TEMPLATE": 3,
}


def _wave1_pids() -> set[int]:
    if not WAVE1_RESULTS.exists():
        return set()
    pids: set[int] = set()
    with WAVE1_RESULTS.open() as f:
        for line in f:
            try:
                r = json.loads(line)
                if r.get("pid") is not None:
                    pids.add(int(r["pid"]))
            except Exception:
                continue
    return pids


def main() -> None:
    random.seed(43)  # different seed than wave 1
    OUT.mkdir(parents=True, exist_ok=True)

    wave1_done = _wave1_pids()
    print(f"Wave 1 already-probed pids: {len(wave1_done)}")

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Properties"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}

    pool: dict[str, list[dict]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: row[i] for i, h in enumerate(hdr)}
        tier = rec.get("extraction_tier") or "?"
        if tier in SKIP_TIERS:
            continue
        pid = rec.get("apartment_id")
        if pid in wave1_done:
            continue
        n_full = rec.get("n_full_post_fix") or 0
        if n_full == 0:
            pool[tier].append(rec)

    print(f"\nTotal n_full=0 props (excl. wave 1 + shipped tiers): {sum(len(p) for p in pool.values())}")

    sample = []
    for tier, quota in N_FULL_ZERO_QUOTA.items():
        p = pool.get(tier, [])
        if not p:
            print(f"  WARN  no props for {tier}")
            continue
        picked = random.sample(p, min(quota, len(p)))
        for r in picked:
            r["_probe_cohort"] = "n_full_zero_w2"
            r["_probe_tier"] = tier
            r["_pool_size"] = len(p)
            sample.append(r)
        print(f"  {len(picked):>2}/{quota} from {tier} (pool={len(p)})")

    out_path = OUT / "n_full_zero_w2_worklist.jsonl"
    with out_path.open("w") as f:
        for r in sample:
            f.write(json.dumps(r, default=str) + "\n")
    print(f"\nWROTE {out_path}: {len(sample)} props")


if __name__ == "__main__":
    main()
