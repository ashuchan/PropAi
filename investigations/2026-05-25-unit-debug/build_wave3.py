"""Wave 3 worklist — 200 fresh props from remaining n_full=0 cohort.

Excludes:
- Wave 1 + Wave 2 + sqft=-1 sample (already probed: ~292 distinct props)
- Tiers owned by shipped chips (Entrata PP, Cobblestone EdificeCMS,
  ChocolateWorks Reinhold, PRG FortressTech, parsing regex bundle)
"""
from __future__ import annotations

import json
import random
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

XLSX = "/Users/ankur/Downloads/canary_1ef1060_2026-05-25_post_phase16_v2.xlsx"
OUT = Path(__file__).parent / "artifacts" / "probe"


def _done_pids() -> set[int]:
    pids: set[int] = set()
    for fname in ("n_full_zero_results.jsonl", "n_full_zero_w2_results.jsonl",
                  "low_strict_results.jsonl", "sqft_neg1_results.jsonl"):
        p = OUT / fname
        if not p.exists():
            continue
        with p.open() as f:
            for line in f:
                try:
                    pids.add(int(json.loads(line).get("pid")))
                except Exception:
                    pass
    return pids


def main() -> None:
    random.seed(45)
    OUT.mkdir(parents=True, exist_ok=True)

    done = _done_pids()
    print(f"Already-probed pids across all waves: {len(done)}")

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Properties"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}

    SKIP_TIERS = {
        "TIER_1_DOM_ENTRATA_PP_SSR_PLAN_LEVEL",
        "TIER_1_API_RENTMANAGER",
        "TIER_1_API_REPLI360_PLAN_LEVEL",
        # ChocolateWorks/Reinhold cohort already shipped
        # (no specific tier — Cluster ID was rr-unit-block DOM)
    }

    pool: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[idx["apartment_id"]]
        if pid in done:
            continue
        tier = row[idx["extraction_tier"]] or "?"
        if tier in SKIP_TIERS:
            continue
        n_full = row[idx["n_full_post_fix"]] or 0
        if n_full == 0:
            rec = {h: row[i] for i, h in enumerate(hdr)}
            rec["_probe_cohort"] = "n_full_zero_w3"
            rec["_probe_tier"] = tier
            pool.append(rec)

    print(f"Eligible remaining n_full=0 pool: {len(pool)}")
    random.shuffle(pool)
    sample = pool[:200]
    print(f"Sampling: {len(sample)}")

    # Tier breakdown
    from collections import Counter
    tc = Counter(r["_probe_tier"] for r in sample)
    print("\n--- Wave 3 sample tier mix ---")
    for t, c in tc.most_common(15):
        print(f"  {c:>3}  {t}")

    out_path = OUT / "n_full_zero_w3_worklist.jsonl"
    with out_path.open("w") as f:
        for r in sample:
            f.write(json.dumps(r, default=str) + "\n")
    print(f"\nWROTE {out_path}: {len(sample)} props")


if __name__ == "__main__":
    main()
