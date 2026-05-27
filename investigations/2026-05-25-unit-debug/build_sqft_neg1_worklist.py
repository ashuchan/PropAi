"""Build stratified sample of properties with sqft=-1 units.

3,483 unit-rows across ~370 distinct properties have area_sqft=-1.
This signals either (a) operator-data-gap (sqft genuinely not published),
or (b) extraction-miss (sqft IS published, parser missed it). We need
to deep-probe a sample to triage.

Stratification:
- One sample of ~50 props weighted by tier pool size
- Skip tiers already owned by completed chips (Cobblestone EdificeCMS,
  Entrata PP per-plan, etc.) because those props are likely getting
  re-extracted next canary anyway.
"""
from __future__ import annotations

import json
import random
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

XLSX = "/Users/ankur/Downloads/canary_1ef1060_2026-05-25_post_phase16_v2.xlsx"
OUT = Path(__file__).parent / "artifacts" / "probe"

# Per-tier sample quotas — total ~50 props
# Skipping tiers owned by completed chips (Entrata PP, Cobblestone)
QUOTA = {
    "TIER_1_DOM_APPFOLIO_VANITY": 8,        # 95-prop pool — biggest cohort
    "TIER_1_DOM_GENERIC_PLAN_TEXT": 8,      # 96-prop pool (already lifted by Elementor fix; sample rest)
    "TIER_1_API_RENTCAFE_SECURECAFE": 6,    # 52-prop pool
    "TIER_MERGED_CROSS_PAGE": 5,            # 34-prop pool (Walnut Creek lineage)
    "TIER_1_DOM_APPFOLIO_VANITY_PLAN_LEVEL": 3,  # 9-prop pool
    "TIER_3_DOM": 5,                        # 61-prop pool
    "TIER_1_DOM_GENERIC_PLAN_TEXT_PLAN_LEVEL": 4,  # 38-prop pool
    "TIER_1_API_REPLI360_PLAN_LEVEL": 2,    # 3-prop pool — small
    "TIER_1_API": 4,                         # 9-prop pool
    "TIER_1_API_REPLI360": 3,                # 10-prop pool
    "TIER_1_API_ONESITE_WORKFLOW": 2,        # 5-prop pool
    "TIER_1_API_SIGHTMAP_IFRAME": 2,         # 10-prop pool
}


def main() -> None:
    random.seed(44)
    OUT.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Units"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr)}

    # Collect props with sqft=-1, by tier; track count of -1 units per prop
    by_tier_props: dict[str, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    prop_info: dict[int, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["area_sqft"]] != -1:
            continue
        tier = row[idx["extraction_tier"]] or "?"
        pid = row[idx["apartment_id"]]
        by_tier_props[tier][pid] += 1
        if pid not in prop_info:
            prop_info[pid] = {
                "apartment_id": pid,
                "name": row[idx["name"]],
                "tier_observed": tier,
            }

    # Load Properties sheet for URL + address
    ws_props = wb["Properties"]
    phdr = [c.value for c in next(ws_props.iter_rows(min_row=1, max_row=1))]
    pidx = {h: i for i, h in enumerate(phdr)}
    for row in ws_props.iter_rows(min_row=2, values_only=True):
        pid = row[pidx["apartment_id"]]
        if pid in prop_info:
            for k in ("website", "address", "city", "state", "zip"):
                if k in pidx:
                    prop_info[pid][k] = row[pidx[k]]

    # Sample per tier
    sample = []
    for tier, quota in QUOTA.items():
        pool = list(by_tier_props.get(tier, {}).items())
        if not pool:
            print(f"  WARN no props for {tier}")
            continue
        # Sort pool by neg1_count desc so most-affected props get sampled first
        pool.sort(key=lambda kv: -kv[1])
        # Random sample from top half (more interesting cases)
        top_half = pool[:max(quota * 3, 10)]
        picked = random.sample(top_half, min(quota, len(top_half)))
        for pid, neg1_count in picked:
            rec = dict(prop_info[pid])
            rec["_probe_cohort"] = "sqft_neg1"
            rec["_probe_tier"] = tier
            rec["_pool_size"] = len(pool)
            rec["_neg1_unit_count"] = neg1_count
            sample.append(rec)
        print(f"  {len(picked):>2}/{quota} from {tier} (pool={len(pool)})")

    out_path = OUT / "sqft_neg1_worklist.jsonl"
    with out_path.open("w") as f:
        for r in sample:
            f.write(json.dumps(r, default=str) + "\n")
    print(f"\nWROTE {out_path}: {len(sample)} props")


if __name__ == "__main__":
    main()
