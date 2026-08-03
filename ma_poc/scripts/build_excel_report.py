#!/usr/bin/env python3
"""Build a shareable Excel workbook from a Jugnu run's output.

A stopgap for the (not-yet-hosted) dashboard: reads ``properties.json`` (and
``report.json`` if present) from a run directory and writes ``report.xlsx``
with three sheets:

  * **Summary**    — run totals, verdict / tier distribution, field-fill rates,
                     data-quality mix (from report.json, else recomputed).
  * **Properties** — one row per property: identity + verdict + provenance
                     (confidence, adapter, winning tier, fetch method,
                     data-quality counts).
  * **Units**      — one row per unit, flattened, incl. per-unit
                     ``extraction_tier`` and ``is_floor_plan_level``.

Defensive: works on the v2 output shape and tolerates older runs that predate
the provenance surfacing (PR #95) — missing fields render blank.

Usage:
    python -m ma_poc.scripts.build_excel_report --run-dir data/v2/runs/2026-07-12
    python -m ma_poc.scripts.build_excel_report --properties props.json --out out.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ma_poc.core.schema_v2 import field_is_absent

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13)
_SUBTITLE_FONT = Font(bold=True, size=11, color="1F3864")

_FILL_FIELDS = (
    "unit_id",
    "beds",
    "baths",
    "area",
    "rent_low",
    "available_date",
    "availability_status",
    "floor_plan_name",
    "floor",
    "building",
    "source_unit_id",
    "canonical_unit_id",
    "building_id",
    "unit_history_key",
    "availability_date_provenance",
    "concession_text",
    "source_ids",
)

_UNAVAILABLE_TOKENS = {
    "unavailable",
    "not available",
    "rented",
    "leased",
    "leased out",
    "rented out",
    "occupied",
    "pending",
    "coming soon",
    "off market",
    "off-market",
}


# ── field extraction (defensive across v1/v2 + provenance-present-or-absent) ──


def _first(d: dict[str, Any], *keys: str) -> Any:
    for k in keys:
        v = d.get(k)
        if v not in (None, "", "null"):
            return v
    return None


def _prop_id(p: dict[str, Any]) -> str:
    meta = p.get("_meta") or {}
    return str(_first(p, "apartment_id", "Property ID", "Unique ID") or meta.get("canonical_id") or "")


def _prop_name(p: dict[str, Any]) -> str:
    return str(_first(p, "proj_name", "Property Name", "name") or "")


def _property_row(p: dict[str, Any]) -> dict[str, Any]:
    meta = p.get("_meta") or {}
    prov = meta.get("provenance") or {}
    ex = p.get("_extract_result") or {}
    fetch = prov.get("fetch") or {}
    dq = prov.get("data_quality") or {}
    units = p.get("units") or []
    return {
        "property_id": _prop_id(p),
        "name": _prop_name(p),
        "city": _first(p, "city", "City") or "",
        "state": _first(p, "state", "State") or "",
        "zip": _first(p, "zip_code", "ZIP Code", "zip") or "",
        "website": _first(p, "website", "Website") or "",
        "verdict": meta.get("verdict") or "",
        "verdict_reason": meta.get("verdict_reason") or "",
        "units": len(units),
        "confidence": prov.get("confidence"),
        "adapter": prov.get("adapter") or "",
        "detected_pms": prov.get("detected_pms") or "",
        "winning_tier": prov.get("winning_tier")
        or (ex.get("tier_used") if isinstance(ex, dict) else "")
        or "",
        "fetch_outcome": fetch.get("outcome") or "",
        "render_mode": fetch.get("render_mode") or "",
        "proxied": fetch.get("proxied"),
        "page_load_ms": fetch.get("page_load_ms"),
        "real_id_units": dq.get("real_id_units"),
        "synthetic_id_units": dq.get("synthetic_id_units"),
        "plan_level_units": dq.get("plan_level_units"),
        "is_lease_up": prov.get("is_lease_up"),
        "llm_cost_usd": ex.get("llm_cost_usd") if isinstance(ex, dict) else None,
        "concession_banner": _concession_text(p.get("concessions")),
    }


def _concession_text(c: Any) -> str:
    if not c:
        return ""
    if isinstance(c, str):
        return c
    if isinstance(c, dict):
        return str(c.get("text") or c.get("banner") or c.get("value") or "")
    return str(c)


def _scraped_date(p: dict[str, Any], report: dict[str, Any] | None, u: dict[str, Any] | None = None) -> str:
    # Prefer run-level date when available; fallback to the unit's capture date
    # so this remains useful outside a run-report artifact.
    if report and report.get("run_date"):
        return str(report.get("run_date"))
    if u:
        dc = u.get("date_captured")
        if dc:
            return str(dc).split()[0]
    for unit in _property_report_rows(p):
        if not isinstance(unit, dict):
            continue
        dc = unit.get("date_captured")
        if dc:
            return str(dc).split()[0]
    return ""


def _property_url(p: dict[str, Any]) -> str:
    return str(_first(p, "website", "Website", "url", "URL") or "")


def _is_available_unit_for_report(u: dict[str, Any]) -> bool:
    status = _first(u, "availability_status", "availability")
    if status is None:
        return True
    s = str(status).strip().lower()
    if not s:
        return True
    if s in _UNAVAILABLE_TOKENS:
        return False
    return not any(token in s for token in _UNAVAILABLE_TOKENS)


def _property_report_rows(p: dict[str, Any]) -> list[dict[str, Any]]:
    """Return UNIT rows plus every canonical ``floor_plans[]`` row.

    Plan rows are copied and explicitly marked; they never acquire a synthetic
    apartment identity.  A legacy artifact may already have a plan row in
    ``units[]`` as well as the canonical array, so a conservative plan key
    prevents double export while retaining distinct plans.
    """

    rows: list[dict[str, Any]] = []
    seen_plans: set[tuple[str, str, str, str, str]] = set()
    for raw in p.get("units") or []:
        if not isinstance(raw, dict):
            continue
        row = dict(raw)
        rows.append(row)
        if row.get("is_floor_plan_level"):
            seen_plans.add(
                tuple(
                    str(row.get(key) or "").strip().casefold()
                    for key in ("floor_plan_id", "floor_plan_name", "beds", "baths", "area_sqft")
                )
            )
    for raw in p.get("floor_plans") or []:
        if not isinstance(raw, dict):
            continue
        row = dict(raw)
        row["is_floor_plan_level"] = True
        key = tuple(
            str(row.get(field) or "").strip().casefold()
            for field in ("floor_plan_id", "floor_plan_name", "beds", "baths", "area_sqft")
        )
        if key in seen_plans:
            continue
        seen_plans.add(key)
        rows.append(row)
    return rows


def _rp_public_unit_id(u: dict[str, Any]) -> str:
    """Return the operator-visible unit label for the RP presentation layer.

    ``unit_id`` remains the collision-safe storage identity.  Knock is the
    important legacy fallback: older formatted artifacts retained the public
    label only in ``unit_name`` while putting ``knock_unit_id-<uuid>`` in the
    canonical field.
    """
    public = _first(u, "unit_number", "display_unit_id", "marketing_unit_number")
    if public not in (None, "", "null"):
        return str(public).strip()
    canonical = str(u.get("unit_id") or "").strip()
    if canonical.startswith("knock_unit_id-"):
        label = str(u.get("unit_name") or "").strip()
        if label:
            return label
    return canonical


def _unit_row(p: dict[str, Any], u: dict[str, Any]) -> dict[str, Any]:
    sids = u.get("source_ids") or {}
    area_sqft = u.get("area_sqft")
    if "area_sqft" not in u:
        legacy_area = _first(u, "area", "sqft")
        area_sqft = None if legacy_area == -1 else legacy_area
    return {
        "property_id": _prop_id(p),
        "property_name": _prop_name(p),
        "unit_id": u.get("unit_id") or "",
        "source_unit_id": u.get("source_unit_id") or "",
        "canonical_unit_id": u.get("canonical_unit_id") or u.get("unit_id") or "",
        "unit_history_key": u.get("unit_history_key") or "",
        "unit_history_key_quality": u.get("unit_history_key_quality") or "",
        "unit_history_key_version": u.get("unit_history_key_version") or "",
        "is_floor_plan_level": u.get("is_floor_plan_level"),
        "extraction_tier": u.get("extraction_tier") or "",
        "beds": u.get("beds"),
        "baths": u.get("baths"),
        "area_sqft": area_sqft,
        "area_is_published": u.get("area_is_published"),
        "area_low": u.get("area_low"),
        "area_high": u.get("area_high"),
        "area_range": u.get("area_range") or "",
        "area_range_raw": u.get("area_range_raw") or "",
        "area_value_type": u.get("area_value_type") or "",
        "area_provenance": u.get("area_provenance") or "",
        "area_source_url": u.get("area_source_url") or "",
        "rent_low": u.get("rent_low") or u.get("market_rent_low"),
        "rent_high": u.get("rent_high") or u.get("market_rent_high"),
        "rent_range": u.get("rent_range") or "",
        "rent_range_raw": u.get("rent_range_raw") or u.get("_rent_range_raw") or "",
        "rent_is_range": u.get("rent_is_range"),
        "rent_provenance": u.get("rent_provenance") or "",
        "availability_status": u.get("availability_status") or "",
        "available_date": u.get("available_date") or "",
        "available_date_raw": u.get("available_date_raw") or u.get("_available_date_raw") or "",
        "availability_date_provenance": u.get("availability_date_provenance") or "",
        "floor_plan_name": u.get("floor_plan_name") or "",
        "floor_plan_name_provenance": u.get("floor_plan_name_provenance") or "",
        "floor_plan_id": u.get("floor_plan_id") or "",
        "floor": u.get("floor"),
        "building": u.get("building") or "",
        "building_id": u.get("building_id") or "",
        "building_id_source": u.get("building_id_source") or "",
        "lease_term": u.get("lease_term") or "",
        "move_in_date": u.get("move_in_date") or "",
        "concession": u.get("concession_text_clean") or u.get("concession_text") or "",
        "source_ids": json.dumps(sids) if sids else "",
        "identity_quality": u.get("identity_quality") or "",
        "unit_id_aliases": json.dumps(u.get("unit_id_aliases") or []),
        "source_response_sha256": u.get("source_response_sha256") or "",
        "source_response_url": u.get("source_response_url") or "",
        "source_record_locator": u.get("source_record_locator") or "",
        "source_asset_url": u.get("source_asset_url") or "",
        "source_asset_sha256": u.get("source_asset_sha256") or "",
    }


def _rp_row(
    p: dict[str, Any],
    u: dict[str, Any],
    report: dict[str, Any] | None,
    plan_no_by_property: dict[str, dict[str, int]],
) -> dict[str, Any]:
    property_id = _prop_id(p)
    plan_key = str(
        _first(
            u,
            "floor_plan_id",
            "floor_plan_name",
            "floor_plan_type",
            "floorplan_name",
            "floorplanname",
            "floorplan",
        )
        or ""
    )
    if not plan_key:
        if u.get("is_floor_plan_level"):
            plan_key = "plan-level"
        else:
            plan_key = "unit-level"
    by_property = plan_no_by_property.setdefault(property_id, {})
    if plan_key in by_property:
        plan_no = by_property[plan_key]
    else:
        plan_no = len(by_property) + 1
        by_property[plan_key] = plan_no

    return {
        "apartmentid": property_id,
        "scrapeddate": _scraped_date(p, report, u),
        "floorplannumber": plan_no,
        "floorplanname": u.get("floor_plan_name") or "",
        "unit_or_plan_level": "PLAN" if u.get("is_floor_plan_level") else "UNIT",
        "area": _first(u, "area_sqft", "area") if u.get("area_sqft") is not None else None,
        "beds": u.get("beds"),
        "baths": u.get("baths"),
        # A floor plan is not an apartment.  Keep the RP field blank instead
        # of manufacturing a unit identity from its plan key.
        "unitid": "" if u.get("is_floor_plan_level") else _rp_public_unit_id(u),
        "marketrentlow": u.get("rent_low") if u.get("rent_low") is not None else u.get("market_rent_low"),
        "marketrenthigh": u.get("rent_high") if u.get("rent_high") is not None else u.get("market_rent_high"),
        "availabledate": u.get("available_date") or u.get("available_date_raw") or "",
        "property_name": _prop_name(p),
        "address": _first(p, "address", "Address") or "",
        "url": _property_url(p),
    }


# ── sheet writers ────────────────────────────────────────────────────────────


def _write_table(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]], start_row: int = 1) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=start_row + 1):
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r, column=c, value=row.get(h))
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(rows)}"
    _autosize(ws, headers, rows)


def _autosize(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    for c, h in enumerate(headers, start=1):
        width = len(str(h))
        for row in rows[:200]:  # sample for speed on huge sheets
            v = row.get(h)
            if v is not None:
                width = max(width, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 8), 60)


def _write_summary(ws: Worksheet, report: dict[str, Any], properties: list[dict[str, Any]]) -> None:
    row = 1

    def put(label: str, value: Any, *, title: bool = False, sub: bool = False) -> None:
        nonlocal row
        cell = ws.cell(row=row, column=1, value=label)
        if title:
            cell.font = _TITLE_FONT
        elif sub:
            cell.font = _SUBTITLE_FONT
        else:
            cell.font = Font(bold=True)
        if value is not None:
            ws.cell(row=row, column=2, value=value)
        row += 1

    totals = report.get("totals") or {}
    put(f"Run Report — {report.get('run_date', '')}", None, title=True)
    row += 1
    put("Properties", totals.get("properties", len(properties)))
    put("Succeeded", totals.get("succeeded"))
    put("Failed", totals.get("failed"))
    put("Success rate %", totals.get("success_rate_pct"))
    put("Carry-forward", totals.get("carry_forward"))
    put("Operator-transparency", totals.get("operator_transparency"))
    row += 1

    def put_counter(title: str, d: dict[str, Any]) -> None:
        nonlocal row
        put(title, None, sub=True)
        for k, v in (d or {}).items():
            ws.cell(row=row, column=1, value=str(k))
            ws.cell(row=row, column=2, value=v)
            row += 1
        row += 1

    put_counter("Verdict distribution", report.get("verdict_distribution", {}))
    put_counter("Tier distribution", report.get("tier_distribution", {}))

    dq = report.get("data_quality") or {}
    put("Data quality", None, sub=True)
    put("  total units", dq.get("total_units"))
    put("  real-id units", dq.get("real_id_units"))
    put("  synthetic-id units", dq.get("synthetic_id_units"))
    put("  plan-level units", dq.get("plan_level_units"))
    put_counter("  by tier family", dq.get("by_tier_family", {}))

    put_counter("Field fill rates (%)", report.get("field_fill_rates", {}))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40


def _recompute_report(properties: list[dict[str, Any]]) -> dict[str, Any]:
    """Minimal summary when report.json is absent — verdicts + units + fills."""
    verdicts: Counter[str] = Counter()
    tiers: Counter[str] = Counter()
    fill = dict.fromkeys(_FILL_FIELDS, 0)
    n_units = 0
    for p in properties:
        verdicts[str((p.get("_meta") or {}).get("verdict") or "UNKNOWN")] += 1
        for u in _property_report_rows(p):
            n_units += 1
            tiers[str(u.get("extraction_tier") or "NONE")] += 1
            for k in _FILL_FIELDS:
                # Same sentinel rule as run_report.build — ``area`` uses the
                # -1 ABSENT sentinel, which an inline emptiness test scores
                # as filled. See schema_v2.field_is_absent.
                if not field_is_absent(k, u.get(k)):
                    fill[k] += 1
    succeeded = sum(v for k, v in verdicts.items() if k.startswith("SUCCESS"))
    total = len(properties)
    return {
        "run_date": "",
        "totals": {
            "properties": total,
            "succeeded": succeeded,
            "failed": total - succeeded,
            "success_rate_pct": round(100 * succeeded / total, 2) if total else 0,
        },
        "verdict_distribution": dict(verdicts.most_common()),
        "tier_distribution": dict(tiers.most_common()),
        "data_quality": {"total_units": n_units},
        "field_fill_rates": {
            k: (round(100 * fill[k] / n_units, 1) if n_units else 0.0) for k in _FILL_FIELDS
        },
    }


# ── workbook assembly ────────────────────────────────────────────────────────

_PROP_HEADERS = [
    "property_id",
    "name",
    "city",
    "state",
    "zip",
    "website",
    "verdict",
    "verdict_reason",
    "units",
    "confidence",
    "adapter",
    "detected_pms",
    "winning_tier",
    "fetch_outcome",
    "render_mode",
    "proxied",
    "page_load_ms",
    "real_id_units",
    "synthetic_id_units",
    "plan_level_units",
    "is_lease_up",
    "llm_cost_usd",
    "concession_banner",
]
_UNIT_HEADERS = [
    "property_id",
    "property_name",
    "unit_id",
    "source_unit_id",
    "canonical_unit_id",
    "unit_history_key",
    "unit_history_key_quality",
    "unit_history_key_version",
    "is_floor_plan_level",
    "extraction_tier",
    "beds",
    "baths",
    "area_sqft",
    "area_is_published",
    "area_low",
    "area_high",
    "area_range",
    "area_range_raw",
    "area_value_type",
    "area_provenance",
    "area_source_url",
    "rent_low",
    "rent_high",
    "rent_range",
    "rent_range_raw",
    "rent_is_range",
    "rent_provenance",
    "availability_status",
    "available_date",
    "available_date_raw",
    "availability_date_provenance",
    "floor_plan_name",
    "floor_plan_name_provenance",
    "floor_plan_id",
    "floor",
    "building",
    "building_id",
    "building_id_source",
    "lease_term",
    "move_in_date",
    "concession",
    "source_ids",
    "identity_quality",
    "unit_id_aliases",
    "source_response_sha256",
    "source_response_url",
    "source_record_locator",
    "source_asset_url",
    "source_asset_sha256",
]
_RP_HEADERS = [
    "apartmentid",
    "scrapeddate",
    "floorplannumber",
    "floorplanname",
    "unit_or_plan_level",
    "area",
    "beds",
    "baths",
    "unitid",
    "marketrentlow",
    "marketrenthigh",
    "availabledate",
    "property_name",
    "address",
    "url",
]


def build_workbook(properties: list[dict[str, Any]], report: dict[str, Any] | None) -> Workbook:
    """Assemble the 4-sheet workbook. Pure; safe on empty input."""
    rep = report or _recompute_report(properties)
    wb = Workbook()
    _write_summary(wb.active, rep, properties)  # type: ignore[arg-type]
    wb.active.title = "Summary"

    ws_p = wb.create_sheet("Properties")
    _write_table(ws_p, _PROP_HEADERS, [_property_row(p) for p in properties])

    ws_u = wb.create_sheet("Units")
    unit_rows = [_unit_row(p, u) for p in properties for u in _property_report_rows(p)]
    _write_table(ws_u, _UNIT_HEADERS, unit_rows)

    ws_rp = wb.create_sheet("RP_Format")
    rp_plan_numbers: dict[str, dict[str, int]] = {}
    rp_rows = [
        _rp_row(p, u, rep if isinstance(rep, dict) else None, rp_plan_numbers)
        for p in properties
        for u in _property_report_rows(p)
        # Plan-level success must be represented even when the operator says
        # waitlist/contact/not available.  Availability filtering applies only
        # to physical UNIT rows.
        if u.get("is_floor_plan_level") or _is_available_unit_for_report(u)
    ]
    _write_table(ws_rp, _RP_HEADERS, rp_rows)
    return wb


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Build an Excel report from a Jugnu run.")
    ap.add_argument("--run-dir", type=Path, help="Run dir containing properties.json + report.json")
    ap.add_argument("--properties", type=Path, help="Path to properties.json (overrides --run-dir)")
    ap.add_argument("--report", type=Path, help="Path to report.json (optional)")
    ap.add_argument("--out", type=Path, help="Output .xlsx (default: <run-dir>/report.xlsx)")
    args = ap.parse_args(argv)

    if args.properties:
        props_path = args.properties
        report_path = args.report
        out = args.out or props_path.with_name("report.xlsx")
    elif args.run_dir:
        props_path = args.run_dir / "properties.json"
        report_path = args.report or (args.run_dir / "report.json")
        out = args.out or (args.run_dir / "report.xlsx")
    else:
        ap.error("provide --run-dir or --properties")
        return 2

    if not props_path.exists():
        print(f"error: {props_path} not found", file=sys.stderr)
        return 1
    properties = json.loads(props_path.read_text(encoding="utf-8"))
    if isinstance(properties, dict):
        properties = properties.get("properties") or list(properties.values())
    report = None
    if report_path and report_path.exists():
        try:
            report = json.loads(report_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            report = None

    wb = build_workbook(properties, report)
    wb.save(out)
    n_rows = sum(len(_property_report_rows(p)) for p in properties)
    print(f"wrote {out}  ({len(properties)} properties, {n_rows} unit/plan rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
