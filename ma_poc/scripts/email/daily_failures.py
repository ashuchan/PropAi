"""Email a daily PropAi health & failures report with two XLSX attachments.

Companion to ``email_daily_report.py`` — that script ships the headline
dashboard summary; this one ships the operational drill-down: every
property + unit scraped today, plus every failed property with the
debug breadcrumbs needed to triage it (latest scrape-event,
``failure_reason``, error codes from ``run_issues``).

Two attachments, both ``.xlsx`` (Gmail previews them as Sheets natively):

    1. ``scraped_units_{run_date}.xlsx``
       One row per unit captured today, joined with property metadata
       (canonical_id, name, city, state, website, run-level verdict).
       Sourced from ``property_snapshots.payload`` so it includes
       carry-forward units and reflects the exact unit list the runner
       persisted for the day.

    2. ``failed_properties_{run_date}.xlsx``
       One row per failed canonical_id (``run_ledger.status`` not in
       ``reporting.verdict._SUCCESS_VERDICTS`` — i.e. excluding SUCCESS /
       SUCCESS_PARTIAL / SUCCESS_PLAN_LEVEL) with: error/warning counts,
       every issue code from ``run_issues``
       concatenated, and the latest ``scrape_events`` row's
       ``scrape_outcome`` + ``failure_reason``. Designed so a person
       sorting by error_code or by failure_reason text can immediately
       see what broke and which properties broke the same way.

Data sources
------------

The runner writes per-shard artifacts to two destinations on every run:

  • Cloud SQL — ``sync.run_to_pg`` mirrors snapshots, ledger, issues,
    scrape_events into the proppy/jugnu instance. The script's primary
    read path. Subject to the 3-day retention sweep
    (``ma_poc/scripts/sync/run_to_pg.py::_apply_retention``) — anything
    older than 3 days has been deleted from per-run tables.

  • GCS — ``runners/shard_entry.py`` uploads the entire
    ``runs/{run_date}/`` tree per shard to
    ``gs://{BUCKET_NAME}/runs/{run_date}/shard_{idx}/``. Kept
    indefinitely; this is the long-window source for any date that has
    aged out of Cloud SQL.

Resolution order:

  1. If ``--run-date`` is given, use it verbatim.
  2. Otherwise read ``runs.list_runs()`` from Cloud SQL and pick the
     newest entry. If today (UTC) is in the list, prefer it.
  3. If Cloud SQL has no rows for the chosen date AND ``BUCKET_NAME``
     is set, fall through to downloading
     ``gs://{BUCKET_NAME}/runs/{run_date}/shard_*/properties.json``
     and ``issues.jsonl`` and rebuild the report from those instead.

Both signals are surfaced in the email body so the reader can see at a
glance whether SQL retention has trimmed the run already.

Usage
-----

    # Today's run (UTC), email to REPORT_RECIPIENTS
    python scripts/email_daily_failures_report.py

    # Specific run
    python scripts/email_daily_failures_report.py --run-date 2026-05-07

    # Override recipients
    python scripts/email_daily_failures_report.py --recipients a@x.com,b@y.com

    # Render only — write XLSX + print HTML to stdout, do not send
    python scripts/email_daily_failures_report.py --dry-run

    # Pin the output dir (defaults to ma_poc/data/runs/{run_date}/)
    python scripts/email_daily_failures_report.py --out-dir C:\\tmp\\reports

    # Force the GCS fallback even when SQL has rows (post-retention test)
    python scripts/email_daily_failures_report.py --use-gcs

Required env (already in ma_poc/.env):

    REPORT_RECIPIENTS       Comma-separated default recipient list
    REPORT_SENDER_NAME      Display name in the email From header
    EMAIL_TRANSPORT         gmail_api (default) | mcp — see
                            ``scripts/email/daily.py`` for the per-transport
                            env vars (GMAIL_DELEGATED_USER for gmail_api;
                            GMAIL_MCP_COMMAND/ARGS for mcp).
    DATA_PROVIDER           postgres | sqlite | filesystem
    DATABASE_URL            postgresql+pg8000://user:pass@host:port/dbname
    CLOUD_SQL_INSTANCE      project:region:instance (Cloud SQL path)
    BUCKET_NAME             gs://{BUCKET_NAME}/runs/{date}/shard_*/ — used
                            for the GCS fallback path. Without it the
                            script still works against Cloud SQL only.

Windows Task Scheduler — add / remove
-------------------------------------

For daily scheduling, point Task Scheduler at the wrapper batch file
``scripts/send_daily_failures_report.bat`` rather than this script
directly. The wrapper handles venv activation, working-directory
quirks, and an optional ``--with-sync`` step that mirrors Cloud SQL
into the local DB first.

ADD a daily 8:30 AM job (run as the user, NOT elevated):

    schtasks /Create ^
        /TN "PropAi Daily Failures Report" ^
        /TR "C:\\Users\\ashus\\OneDrive\\Documents\\Code\\PropAi\\ma_poc\\scripts\\send_daily_failures_report.bat" ^
        /SC DAILY /ST 08:30 ^
        /F

REMOVE:

    schtasks /Delete /TN "PropAi Daily Failures Report" /F
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import tempfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

_MA_POC_ROOT = Path(__file__).resolve().parent.parent.parent
_REPO_ROOT = _MA_POC_ROOT.parent
for _p in (_MA_POC_ROOT, _REPO_ROOT):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from dotenv import load_dotenv  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from data_provider.sql.models import (  # noqa: E402
    PropertyRow,
    PropertySnapshotRow,
    RunIssueRow,
    RunLedgerRow,
    ScrapeEventRow,
)
from data_provider.sql.provider import SqlDataProvider  # noqa: E402
from ma_poc.core.concession_enrich import Enrichment, enrich_concession  # noqa: E402
from reporting.verdict import _SUCCESS_VERDICTS, verdict_is_success  # noqa: E402
from scripts.email.daily import (  # noqa: E402
    _open_provider,
    _parse_recipients,
)
from scripts.email._client import send_html_email  # noqa: E402

log = logging.getLogger("email_daily_failures_report")


# ── Run-date resolution ──────────────────────────────────────────────────────


def _resolve_run_date(
    provider: SqlDataProvider,
    requested: str | None,
    *,
    bucket: str | None,
) -> tuple[str, bool]:
    """Pick the run_date to report on, plus a flag for GCS presence.

    Resolution order:
      1. ``--run-date`` if supplied.
      2. Today's date (UTC) if the runs table has a row for it.
      3. The newest ``runs.run_date`` we can find.
      4. The newest ``YYYY-MM-DD`` directory under
         ``gs://{bucket}/runs/`` if SQL is empty AND a bucket is given.

    The boolean is True when ``gs://{bucket}/runs/{run_date}/`` has at
    least one shard subdir. Surfaced in the email body for context.
    """
    sql_dates: list[str] = []
    try:
        sql_dates = provider.runs.list_runs()  # ascending YYYY-MM-DD strings
    except Exception as exc:  # noqa: BLE001 — never block on SQL errors
        log.warning("runs.list_runs() failed (%s); falling back to GCS", exc)

    today = datetime.now(timezone.utc).date().isoformat()

    chosen: str | None = None
    if requested:
        chosen = requested
    elif today in sql_dates:
        chosen = today
    elif sql_dates:
        chosen = sql_dates[-1]
        log.warning("No run for %s yet; falling back to latest SQL run %s", today, chosen)

    if chosen is None and bucket:
        gcs_latest = _find_latest_gcs_run(bucket)
        if gcs_latest:
            chosen = gcs_latest
            log.warning("Cloud SQL has no runs; using latest GCS run %s", chosen)

    if chosen is None:
        raise RuntimeError(
            "No runs found in Cloud SQL or GCS. Set BUCKET_NAME and verify "
            "the runner is uploading to gs://{bucket}/runs/{date}/."
        )

    gcs_present = _gcs_run_dir_exists(bucket, chosen) if bucket else False
    return chosen, gcs_present


def _find_latest_gcs_run(bucket: str) -> str | None:
    """List ``gs://{bucket}/runs/`` and return the newest ``YYYY-MM-DD`` prefix."""
    from google.cloud import storage  # local import — only needed on the GCS path

    client = storage.Client()
    iterator = client.list_blobs(bucket, prefix="runs/", delimiter="/")
    # Force iteration so .prefixes is populated.
    for _ in iterator:  # pragma: no cover — empty top-level objects only
        pass
    candidates: list[str] = []
    for prefix in getattr(iterator, "prefixes", []):
        # prefix looks like "runs/2026-05-07/"
        date_str = prefix[len("runs/"):].rstrip("/")
        if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
            candidates.append(date_str)
    return max(candidates) if candidates else None


def _gcs_run_dir_exists(bucket: str | None, run_date: str) -> bool:
    """Return True if ``gs://{bucket}/runs/{run_date}/`` has any shard."""
    if not bucket:
        return False
    from google.cloud import storage

    client = storage.Client()
    iterator = client.list_blobs(
        bucket, prefix=f"runs/{run_date}/", delimiter="/", max_results=1,
    )
    for _ in iterator:
        return True
    return bool(getattr(iterator, "prefixes", []))


# ── Cloud SQL reads ──────────────────────────────────────────────────────────


def _fetch_property_metadata(session: Session, canonical_ids: list[str]) -> dict[str, dict[str, Any]]:
    """Lookup current name/city/state/website per canonical_id.

    The ``properties`` table is upsert-only — it always reflects the
    most recent CSV row + scrape, so a cid that ran today will be
    present even if the CSV column names changed mid-week.
    """
    if not canonical_ids:
        return {}
    rows = session.execute(
        select(PropertyRow).where(PropertyRow.canonical_id.in_(canonical_ids))
    ).scalars()
    return {
        r.canonical_id: {
            "proj_name": r.proj_name or "",
            "city": r.city or "",
            "state": r.state or "",
            "zip_code": r.zip_code or "",
            "website": r.website or "",
            "pmc": r.pmc or "",
            "last_scrape_status": r.last_scrape_status or "",
        }
        for r in rows
    }


def _fetch_scraped_units_from_sql(
    session: Session, run_date: str, meta: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    """Flatten ``property_snapshots.payload`` into one row per unit.

    Each snapshot row has ``payload["units"]`` as a list of unit dicts
    (the V2 schema: unit_id, beds, baths, area, rent_low/high,
    available_date, …) plus ``payload["_meta"]`` carrying ``verdict``
    and ``scrape_tier_used``. We emit one Excel row per unit; properties
    with zero units are emitted as a single placeholder row so the
    sheet still surfaces them.
    """
    rows = session.execute(
        select(PropertySnapshotRow)
        .where(PropertySnapshotRow.run_date == run_date)
        .order_by(PropertySnapshotRow.ordinal.asc())
    ).scalars().all()

    out: list[dict[str, Any]] = []
    for snap in rows:
        cid = snap.canonical_id or ""
        payload = snap.payload or {}
        m = meta.get(cid, {})

        # Property-level fields. Prefer payload (today's scrape) and fall
        # back to the upserted properties row when the snapshot omitted a
        # column (carry-forward records often do).
        prop_name = (
            payload.get("Property Name")
            or payload.get("proj_name")
            or m.get("proj_name", "")
        )
        city = payload.get("City") or m.get("city", "")
        state = payload.get("State") or m.get("state", "")
        zip_code = payload.get("ZIP Code") or m.get("zip_code", "")
        website = payload.get("Website") or m.get("website", "")
        pmc = payload.get("Management Company") or m.get("pmc", "")

        meta_block = payload.get("_meta") or {}
        verdict = meta_block.get("verdict") or ""
        tier = meta_block.get("scrape_tier_used") or ""
        if not tier:
            extract = payload.get("_extract_result") or {}
            tier = extract.get("tier_used") or ""

        # Property-level concession (captured from homepage banner or
        # /specials probe). Per §18 of the playbook the V2 emit puts the
        # property-wide concession at `concessions` / `concessions_clean`
        # / `_concessions_quality`; per-unit `concession_text` is only set
        # when the adapter parsed an explicit per-row offer. When a unit
        # has nothing, fall back to the property banner so xlsx rows for
        # whole-site concessions surface the offer.
        prop_conc_text = payload.get("concessions") or ""
        prop_conc_clean = payload.get("concessions_clean") or ""
        prop_conc_quality = payload.get("_concessions_quality") or ""

        # Run the deterministic enrichment ONCE per property — the
        # property-level banner is reused across every unit row that
        # inherits from it, so we don't pay the regex cost N times.
        prop_enrich_fields = _concession_enrichment_fields(
            prop_conc_text or prop_conc_clean
        )

        units = payload.get("units") or []
        if not units:
            row = {
                "canonical_id": cid,
                "property_name": prop_name,
                "city": city,
                "state": state,
                "zip_code": zip_code,
                "pmc": pmc,
                "website": website,
                "verdict": verdict,
                "tier_used": tier,
                "unit_id": "",
                "floor_plan_name": "",
                "beds": None,
                "baths": None,
                "area": None,
                "rent_low": None,
                "rent_high": None,
                "available_date": "",
                "lease_term": None,
                # Surface the property-level banner even when the property
                # has no units — a no-unit hit is still a place where the
                # operator wants to see whether a concession was advertised.
                "concession_text": _stringify_concessions(prop_conc_text),
                "concession_text_clean": _stringify_concessions(prop_conc_clean),
                "_concession_quality": prop_conc_quality,
                "concessions": _stringify_concessions(
                    prop_conc_clean or prop_conc_text
                ),
            }
            row.update(prop_enrich_fields)
            out.append(row)
            continue

        for u in units:
            # Unit-level fields win when present; otherwise inherit the
            # property-level banner. `concession` (singular) is the legacy
            # per-row key some adapters still emit alongside the V2
            # `concession_text` — keep both in the read chain.
            u_text = u.get("concession_text") or u.get("concession") or ""
            u_clean = u.get("concession_text_clean") or ""
            u_quality = u.get("_concession_quality") or ""
            conc_text = u_text or prop_conc_text
            conc_clean = u_clean or prop_conc_clean
            conc_quality = u_quality or prop_conc_quality
            # When the unit emitted its own per-row concession we re-run
            # the enricher (the producer text differs from the property
            # banner). Otherwise we reuse the property-level enrichment
            # so the per-row cost stays O(1).
            if u_text and u_text != prop_conc_text:
                enrich_fields = _concession_enrichment_fields(u_text)
            else:
                enrich_fields = prop_enrich_fields
            row = {
                "canonical_id": cid,
                "property_name": prop_name,
                "city": city,
                "state": state,
                "zip_code": zip_code,
                "pmc": pmc,
                "website": website,
                "verdict": verdict,
                "tier_used": tier,
                "unit_id": _unit_field(u, "unit_id") or "",
                "floor_plan_name": _unit_field(u, "floor_plan_name") or "",
                "beds": _unit_field(u, "beds"),
                "baths": _unit_field(u, "baths"),
                "area": _unit_field(u, "area"),
                "rent_low": _unit_field(u, "rent_low"),
                "rent_high": _unit_field(u, "rent_high"),
                "available_date": _unit_field(u, "available_date") or "",
                "lease_term": _unit_field(u, "lease_term"),
                "concession_text": _stringify_concessions(conc_text),
                "concession_text_clean": _stringify_concessions(conc_clean),
                "_concession_quality": conc_quality,
                "concessions": _stringify_concessions(
                    conc_clean or conc_text or u.get("concessions")
                ),
            }
            row.update(enrich_fields)
            out.append(row)
    return out


_UNIT_FIELD_SOURCES: dict[str, tuple[str, ...]] = {
    # Read the typed/normalised column first; on null-or-empty, fall back
    # through the producer's raw companions. Mirrors the V2 upsert source
    # map in ``data_provider.sql.stores.SqlUnitStateStore._SNAPSHOT_SOURCES``
    # so the email surfaces whatever the website actually said even when
    # the normaliser couldn't coerce it (e.g. ``available_date`` null but
    # ``available_date_raw="Late August"``).
    "unit_id": ("unit_id", "unit_number", "_unit_number"),
    "floor_plan_name": ("floor_plan_name", "floorplan_name", "_floor_plan"),
    "beds": ("beds", "bedrooms", "_bedrooms"),
    "baths": ("baths", "bathrooms", "_bathrooms"),
    "area": ("area", "sqft", "_sqft"),
    "rent_low": ("rent_low", "market_rent_low", "asking_rent"),
    "rent_high": ("rent_high", "market_rent_high", "asking_rent"),
    "available_date": ("available_date", "available_date_raw", "_date_placeholder"),
    "lease_term": ("lease_term", "_lease_term"),
}


def _unit_field(u: dict[str, Any], field: str) -> Any:
    """Return the first non-empty value for *field* from a unit dict.

    The V2 transform normalises producer values into typed columns, but
    a non-trivial slice of units still ship null on a quality column
    while the producer's raw companion is populated (the canonical
    example is ``available_date`` null + ``available_date_raw=
    "Available 7/24"`` — year missing, can't ISO-normalise). Falling
    back through the chain in ``_UNIT_FIELD_SOURCES`` recovers those
    values without re-deriving them.
    """
    for key in _UNIT_FIELD_SOURCES.get(field, (field,)):
        val = u.get(key)
        if val is None:
            continue
        if isinstance(val, str) and not val.strip():
            continue
        return val
    return None


def _stringify_concessions(value: Any) -> str:
    """Concessions can be free-form text, a dict, or null. Render to one cell."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    except (TypeError, ValueError):
        return str(value)


def _concession_enrichment_fields(raw_text: str) -> dict[str, str]:
    """Run the deterministic enricher on *raw_text* and return the
    xlsx-shaped fields the new columns consume.

    Empty input → empty fields. Never raises.

    The returned dict carries the five enrichment columns:

      * ``concession_banner``      — one-line human summary (~140 chars)
      * ``concession_offer_type``  — canonical taxonomy of the primary
                                     atom (``free_rent`` / ``dollar_off`` / ...)
      * ``concession_target``      — what the discount is APPLIED to
                                     (``rent`` / ``app_fee`` / ``move_in_cost``)
      * ``concession_value``       — short magnitude string (``2 months``,
                                     ``$500``, ``10%``)
      * ``concession_conditions``  — semicolon-joined ``kind:value`` pairs
                                     (``deadline:5/31; lease_length:12+ months;
                                     unit_scope:select``) so a reviewer
                                     can sort/filter by any single signal
    """
    if not raw_text:
        return _empty_enrichment_fields()
    e = enrich_concession(raw_text)
    return _enrichment_to_fields(e)


def _empty_enrichment_fields() -> dict[str, str]:
    return {
        "concession_banner": "",
        "concession_offer_type": "",
        "concession_target": "",
        "concession_value": "",
        "concession_conditions": "",
    }


def _enrichment_to_fields(e: Enrichment) -> dict[str, str]:
    out = _empty_enrichment_fields()
    if e.primary_atom:
        out["concession_offer_type"] = e.primary_atom.offer_type
        out["concession_target"] = e.primary_atom.target
        out["concession_value"] = e.primary_atom.value
    if e.conditions:
        parts = []
        for c in e.conditions:
            if c.value:
                parts.append(f"{c.kind}:{c.value}")
            else:
                parts.append(c.kind)
        out["concession_conditions"] = "; ".join(parts)
    out["concession_banner"] = e.banner or ""
    return out


def _fetch_failed_properties_from_sql(
    session: Session, run_date: str, meta: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    """Build the failure-debug rows.

    Joins three tables in Python (cheaper than a 3-way SQL join given
    the small per-run row counts):
      • ``run_ledger`` — one row per processed property; ``status``
        carries SUCCESS / FAILED / CARRY_FORWARD / etc.
      • ``run_issues`` — many rows per property; we collapse each
        property's issue codes into a comma-separated list.
      • ``scrape_events`` — many rows per property over time; we keep
        the latest timestamp's ``scrape_outcome`` + ``failure_reason``.

    A property is "failed" when its ledger ``status`` is non-empty and
    not exactly ``SUCCESS`` (case-insensitive). CARRY_FORWARD is
    included because the operator usually wants to see those too —
    they're properties that didn't actually succeed today.
    """
    ledger_rows = session.execute(
        select(RunLedgerRow)
        .where(RunLedgerRow.run_date == run_date)
        .order_by(RunLedgerRow.seq.asc())
    ).scalars().all()

    failed_cids: list[str] = []
    ledger_by_cid: dict[str, RunLedgerRow] = {}
    for r in ledger_rows:
        cid = r.canonical_id or ""
        if not cid:
            continue
        status = (r.status or "").upper()
        # Use the canonical success classifier so SUCCESS_PARTIAL,
        # SUCCESS_PLAN_LEVEL etc. don't count as failures here (they're
        # genuine success-class verdicts per reporting.verdict._SUCCESS_VERDICTS).
        if status and not verdict_is_success(status):
            failed_cids.append(cid)
            ledger_by_cid[cid] = r
    if not failed_cids:
        return []

    # Issue codes per cid — keep order and dedupe so the cell stays readable.
    issue_codes_by_cid: dict[str, list[str]] = defaultdict(list)
    issue_messages_by_cid: dict[str, list[str]] = defaultdict(list)
    seen: dict[str, set[str]] = defaultdict(set)
    issue_rows = session.execute(
        select(RunIssueRow)
        .where(RunIssueRow.run_date == run_date)
        .where(RunIssueRow.canonical_id.in_(failed_cids))
        .order_by(RunIssueRow.seq.asc())
    ).scalars()
    for ir in issue_rows:
        cid = ir.canonical_id or ""
        if not cid:
            continue
        code = ir.code or "UNKNOWN"
        if code not in seen[cid]:
            seen[cid].add(code)
            issue_codes_by_cid[cid].append(code)
        if ir.message:
            issue_messages_by_cid[cid].append(_truncate(ir.message, 240))

    # Latest scrape_event per failed cid. property_id mostly == canonical_id
    # in the upserted state, but we fetch by IN and pick the max-timestamp
    # row in Python so the query stays simple.
    event_rows = session.execute(
        select(ScrapeEventRow)
        .where(ScrapeEventRow.property_id.in_(failed_cids))
        .order_by(ScrapeEventRow.scrape_timestamp.desc())
    ).scalars()
    latest_event_by_cid: dict[str, ScrapeEventRow] = {}
    for ev in event_rows:
        # First (newest) wins because we ordered DESC.
        latest_event_by_cid.setdefault(ev.property_id, ev)

    out: list[dict[str, Any]] = []
    for cid in failed_cids:
        ledger = ledger_by_cid[cid]
        ev = latest_event_by_cid.get(cid)
        info = meta.get(cid, {})
        out.append({
            "canonical_id": cid,
            "property_name": info.get("proj_name", ""),
            "city": info.get("city", ""),
            "state": info.get("state", ""),
            "website": ledger.url or info.get("website", ""),
            "pmc": info.get("pmc", ""),
            "ledger_status": (ledger.status or "").upper(),
            "carry_forward_used": bool(ledger.carry_forward_used),
            "scrape_failed": bool(ledger.scrape_failed),
            "units_count": ledger.units_count if ledger.units_count is not None else 0,
            "error_count": ledger.error_count or 0,
            "warning_count": ledger.warning_count or 0,
            "issue_codes": ", ".join(issue_codes_by_cid.get(cid, [])),
            "issue_messages": " | ".join(issue_messages_by_cid.get(cid, [])[:3]),
            "scrape_outcome": (ev.scrape_outcome if ev else "") or "",
            "failure_reason": _truncate((ev.failure_reason if ev else "") or "", 800),
            "extraction_tier": ev.extraction_tier if ev else None,
            "page_load_ms": ev.page_load_ms if ev else None,
            "last_event_at": (ev.scrape_timestamp.isoformat() if ev and ev.scrape_timestamp else ""),
        })
    return out


# ── GCS fallback (used when Cloud SQL has aged the run out) ─────────────────


def _fetch_run_from_gcs(
    bucket: str, run_date: str
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Rebuild scraped-units + failed-properties rows from GCS shards.

    Reads every ``gs://{bucket}/runs/{run_date}/shard_*/properties.json``
    and ``issues.jsonl`` and ``ledger.jsonl``. Returns the same row
    dicts the SQL helpers above produce so the rest of the pipeline
    treats both sources uniformly.

    No ``scrape_events`` data here — the per-shard runner does not
    write events to disk inside the run dir; that table is SQL-only.
    Failure rows from this path therefore have empty
    ``failure_reason`` / ``scrape_outcome``; the email body calls this
    out when the GCS path is used.
    """
    from ma_poc.storage import gcs

    work_dir = Path(tempfile.mkdtemp(prefix=f"failures_{run_date}_"))
    uri_prefix = f"gs://{bucket}/runs/{run_date}/"
    log.info("Downloading %s into %s", uri_prefix, work_dir)
    count = gcs.download_prefix(uri_prefix, work_dir)
    log.info("Downloaded %d objects from %s", count, uri_prefix)

    # Walk every shard subdir.
    scraped: list[dict[str, Any]] = []
    failed_by_cid: dict[str, dict[str, Any]] = {}
    for shard_dir in sorted(work_dir.iterdir()):
        if not shard_dir.is_dir():
            continue
        props_file = shard_dir / "properties.json"
        if props_file.exists():
            scraped.extend(_flatten_properties_json(props_file))
        ledger_file = shard_dir / "ledger.jsonl"
        if ledger_file.exists():
            for ledger in _read_jsonl(ledger_file):
                cid = ledger.get("canonical_id") or ""
                status = (ledger.get("status") or "").upper()
                if cid and status and not verdict_is_success(status):
                    failed_by_cid[cid] = {
                        "canonical_id": cid,
                        "property_name": "",
                        "city": "",
                        "state": "",
                        "website": ledger.get("url", "") or "",
                        "pmc": "",
                        "ledger_status": status,
                        "carry_forward_used": bool(ledger.get("carry_forward_used")),
                        "scrape_failed": bool(ledger.get("scrape_failed")),
                        "units_count": ledger.get("units_count") or 0,
                        "error_count": ledger.get("error_count") or 0,
                        "warning_count": ledger.get("warning_count") or 0,
                        "issue_codes": "",
                        "issue_messages": "",
                        "scrape_outcome": "",
                        "failure_reason": "",
                        "extraction_tier": None,
                        "page_load_ms": None,
                        "last_event_at": "",
                    }
        issues_file = shard_dir / "issues.jsonl"
        if issues_file.exists():
            codes_by_cid: dict[str, list[str]] = defaultdict(list)
            messages_by_cid: dict[str, list[str]] = defaultdict(list)
            seen: dict[str, set[str]] = defaultdict(set)
            for issue in _read_jsonl(issues_file):
                cid = issue.get("canonical_id") or ""
                if not cid:
                    continue
                code = issue.get("code") or "UNKNOWN"
                if code not in seen[cid]:
                    seen[cid].add(code)
                    codes_by_cid[cid].append(code)
                msg = issue.get("message") or ""
                if msg:
                    messages_by_cid[cid].append(_truncate(msg, 240))
            for cid, row in failed_by_cid.items():
                if cid in codes_by_cid:
                    row["issue_codes"] = ", ".join(codes_by_cid[cid])
                if cid in messages_by_cid:
                    row["issue_messages"] = " | ".join(messages_by_cid[cid][:3])

    failed = list(failed_by_cid.values())
    return scraped, failed


def _flatten_properties_json(path: Path) -> list[dict[str, Any]]:
    """Read ``properties.json`` (a list of property dicts) and flatten units.

    Mirrors ``_fetch_scraped_units_from_sql`` so the schema is identical
    no matter which source produced the rows.

    Handles both:
      * V1 legacy keys: ``Unique ID`` / ``Property Name`` / ``City`` / ...
      * V2 internal keys (Jugnu per-shard properties.json): ``apartment_id`` /
        ``proj_name`` / ``city`` / ``state`` / ``zip_code`` / ``pmc`` /
        ``website`` — and ``_meta.canonical_id`` for the canonical id.
    """
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        log.warning("Could not parse %s: %s", path, exc)
        return []
    if not isinstance(data, list):
        return []
    out: list[dict[str, Any]] = []
    for prop in data:
        if not isinstance(prop, dict):
            continue
        meta_block = prop.get("_meta") or {}
        # V2 emits canonical_id only under _meta; V1/CSV legacy emit under
        # "Unique ID" / "canonical_id"; apartment_id is the integer alias
        # the v2 formatter writes at top level.
        cid = (
            meta_block.get("canonical_id")
            or prop.get("Unique ID")
            or prop.get("canonical_id")
            or prop.get("apartment_id")
            or ""
        )
        cid = str(cid) if cid not in (None, "") else ""
        prop_name = (
            prop.get("Property Name")
            or prop.get("proj_name")
            or prop.get("name")
            or ""
        )
        verdict = meta_block.get("verdict") or ""
        tier = meta_block.get("scrape_tier_used") or ""
        if not tier:
            tier = (prop.get("_extract_result") or {}).get("tier_used") or ""

        # Read property-level concessions (V2 emit: `concessions`,
        # `concessions_clean`, `_concessions_quality`, ...). These hold the
        # banner copy captured from the homepage / /specials probe and apply
        # to every unit unless the unit emitted its own per-row concession.
        prop_conc_text = prop.get("concessions") or ""
        prop_conc_clean = prop.get("concessions_clean") or ""
        prop_conc_quality = prop.get("_concessions_quality") or ""

        # Property-level enrichment computed once, reused per unit row
        # that inherits the parent banner.
        prop_enrich_fields = _concession_enrichment_fields(
            prop_conc_text or prop_conc_clean
        )

        common = {
            "canonical_id": cid,
            "property_name": prop_name,
            "city": prop.get("City") or prop.get("city") or "",
            "state": prop.get("State") or prop.get("state") or "",
            "zip_code": prop.get("ZIP Code") or prop.get("zip_code") or "",
            "pmc": prop.get("Management Company") or prop.get("pmc") or "",
            "website": prop.get("Website") or prop.get("website") or "",
            "verdict": verdict,
            "tier_used": tier,
        }
        units = prop.get("units") or []
        if not units:
            placeholder = dict(common)
            placeholder.update({
                "unit_id": "", "floor_plan_name": "",
                "beds": None, "baths": None, "area": None,
                "rent_low": None, "rent_high": None,
                "available_date": "", "lease_term": None,
                # Property-level concession on a unit-less property still
                # belongs in the row so empty-units sites surface their
                # banner.
                "concession_text": _stringify_concessions(prop_conc_text),
                "concession_text_clean": _stringify_concessions(prop_conc_clean),
                "_concession_quality": prop_conc_quality,
                "concessions": _stringify_concessions(
                    prop_conc_clean or prop_conc_text
                ),
            })
            placeholder.update(prop_enrich_fields)
            out.append(placeholder)
            continue
        for u in units:
            row = dict(common)
            # Unit-level concession (V2 emit: `concession_text` etc.) takes
            # precedence; fall back to the property-level banner so xlsx rows
            # for properties with whole-site concessions surface the offer.
            u_text = u.get("concession_text") or u.get("concession") or ""
            u_clean = u.get("concession_text_clean") or ""
            u_quality = u.get("_concession_quality") or ""
            conc_text = u_text or prop_conc_text
            conc_clean = u_clean or prop_conc_clean
            conc_quality = u_quality or prop_conc_quality
            if u_text and u_text != prop_conc_text:
                enrich_fields = _concession_enrichment_fields(u_text)
            else:
                enrich_fields = prop_enrich_fields
            row.update({
                "unit_id": _unit_field(u, "unit_id") or "",
                "floor_plan_name": _unit_field(u, "floor_plan_name") or "",
                "beds": _unit_field(u, "beds"),
                "baths": _unit_field(u, "baths"),
                "area": _unit_field(u, "area"),
                "rent_low": _unit_field(u, "rent_low"),
                "rent_high": _unit_field(u, "rent_high"),
                "available_date": _unit_field(u, "available_date") or "",
                "lease_term": _unit_field(u, "lease_term"),
                "concession_text": _stringify_concessions(conc_text),
                "concession_text_clean": _stringify_concessions(conc_clean),
                "_concession_quality": conc_quality,
                "concessions": _stringify_concessions(
                    conc_clean or conc_text or u.get("concessions")
                ),
            })
            row.update(enrich_fields)
            out.append(row)
    return out


def _fetch_run_from_local_dir(
    run_dir: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Rebuild scraped-units + failed-properties rows from a local mirror.

    Mirrors :func:`_fetch_run_from_gcs` but skips the GCS download step
    — points directly at an already-downloaded run dir such as
    ``C:/tmp/run-2026-05-21/`` with ``shard_*`` subdirectories. The
    layout is identical to what the runner uploads to GCS; see Phase 1
    of ``docs/failed_no_data_debugging_playbook.md``.

    For the failed-rows construction, the local artifacts give us
    ``_meta.verdict`` in ``properties.json`` plus the structured
    ``issues.jsonl``. There is no ``scrape_events`` / ``run_ledger``
    file on disk, so ``scrape_outcome`` / ``failure_reason`` /
    ``page_load_ms`` are derived from ``_meta`` and from the issue
    codes themselves. This is enough for the operator drill-down the
    email targets — the verdict + the issue codes localise the bug
    bucket; the latest ``scrape_events`` row would only add a
    timestamp and a one-line failure reason.
    """
    if not run_dir.exists() or not run_dir.is_dir():
        raise RuntimeError(
            f"Local run directory does not exist or is not a directory: {run_dir}"
        )

    shard_dirs = sorted(
        p for p in run_dir.iterdir()
        if p.is_dir() and p.name.startswith("shard_")
    )
    if not shard_dirs:
        # Sometimes the operator points the flag directly at a shard dir
        # (e.g. a one-property smoke test). Tolerate that by treating the
        # run_dir itself as a single shard.
        if (run_dir / "properties.json").exists():
            shard_dirs = [run_dir]
        else:
            raise RuntimeError(
                f"No shard_* subdirectories under {run_dir} and no "
                f"properties.json directly in it either. Verify the path."
            )

    log.info("Reading %d shard(s) from local mirror %s", len(shard_dirs), run_dir)

    scraped: list[dict[str, Any]] = []
    failed_by_cid: dict[str, dict[str, Any]] = {}

    for shard_dir in shard_dirs:
        props_file = shard_dir / "properties.json"
        if not props_file.exists():
            continue

        # Flatten units (same helper used by the GCS path so V1 + V2 key
        # variants are handled identically — see _flatten_properties_json).
        shard_rows = _flatten_properties_json(props_file)
        scraped.extend(shard_rows)

        # Failed rows: read properties.json directly to consult `_meta`.
        try:
            data = json.loads(props_file.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            log.warning("Could not parse %s for failures: %s", props_file, exc)
            data = []
        if not isinstance(data, list):
            data = []

        for prop in data:
            if not isinstance(prop, dict):
                continue
            meta = prop.get("_meta") or {}
            verdict = (meta.get("verdict") or "").upper()
            if not verdict or verdict_is_success(verdict):
                continue
            cid = (
                str(meta.get("canonical_id") or "")
                or str(prop.get("apartment_id") or "")
                or str(prop.get("Unique ID") or "")
            )
            if not cid:
                continue
            reason = meta.get("verdict_reason") or ""
            errors = meta.get("scrape_errors") or []
            if isinstance(errors, list) and errors:
                reason = reason or "; ".join(str(e) for e in errors[:3])
            failed_by_cid[cid] = {
                "canonical_id": cid,
                "property_name": (
                    prop.get("Property Name")
                    or prop.get("proj_name")
                    or ""
                ),
                "city": prop.get("City") or prop.get("city") or "",
                "state": prop.get("State") or prop.get("state") or "",
                "website": (
                    prop.get("Website") or prop.get("website") or ""
                ),
                "pmc": (
                    prop.get("Management Company") or prop.get("pmc") or ""
                ),
                "ledger_status": verdict,
                "carry_forward_used": bool(meta.get("carry_forward_used")),
                "scrape_failed": True,
                "units_count": len(prop.get("units") or []),
                "error_count": 0,
                "warning_count": 0,
                "issue_codes": "",
                "issue_messages": "",
                # No scrape_events on disk — surface what we have from
                # `_meta` instead so the row isn't empty.
                "scrape_outcome": (
                    meta.get("scrape_tier_used") or ""
                ).upper() or "FAILED",
                "failure_reason": _truncate(reason, 800),
                "extraction_tier": None,
                "page_load_ms": None,
                "last_event_at": "",
            }

        issues_file = shard_dir / "issues.jsonl"
        if issues_file.exists():
            codes_by_cid: dict[str, list[str]] = defaultdict(list)
            messages_by_cid: dict[str, list[str]] = defaultdict(list)
            err_counts: dict[str, int] = defaultdict(int)
            warn_counts: dict[str, int] = defaultdict(int)
            seen: dict[str, set[str]] = defaultdict(set)
            for issue in _read_jsonl(issues_file):
                cid = issue.get("canonical_id") or ""
                if not cid:
                    continue
                sev = (issue.get("severity") or "").upper()
                if sev == "ERROR":
                    err_counts[cid] += 1
                elif sev == "WARNING":
                    warn_counts[cid] += 1
                code = issue.get("code") or "UNKNOWN"
                if code not in seen[cid]:
                    seen[cid].add(code)
                    codes_by_cid[cid].append(code)
                msg = issue.get("message") or ""
                if msg:
                    messages_by_cid[cid].append(_truncate(msg, 240))
            for cid, row in failed_by_cid.items():
                if cid in codes_by_cid:
                    row["issue_codes"] = ", ".join(codes_by_cid[cid])
                if cid in messages_by_cid:
                    row["issue_messages"] = " | ".join(messages_by_cid[cid][:3])
                if cid in err_counts:
                    row["error_count"] = err_counts[cid]
                if cid in warn_counts:
                    row["warning_count"] = warn_counts[cid]

    failed = list(failed_by_cid.values())
    log.info(
        "Local mirror produced %d scraped unit rows, %d failed properties",
        len(scraped), len(failed),
    )
    return scraped, failed


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    """Tolerant JSONL reader — skips blank/malformed lines (truncated files happen)."""
    rows: list[dict[str, Any]] = []
    try:
        with path.open(encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    except OSError as exc:
        log.warning("Could not read %s: %s", path, exc)
    return rows


# ── XLSX writers ────────────────────────────────────────────────────────────


_SCRAPED_COLUMNS: list[tuple[str, str]] = [
    ("canonical_id", "Canonical ID"),
    ("property_name", "Property Name"),
    ("city", "City"),
    ("state", "State"),
    ("zip_code", "ZIP"),
    ("pmc", "Mgmt Company"),
    ("website", "Website"),
    ("verdict", "Verdict"),
    ("tier_used", "Tier Used"),
    ("unit_id", "Unit ID"),
    ("floor_plan_name", "Floor Plan"),
    ("beds", "Beds"),
    ("baths", "Baths"),
    ("area", "Area (sqft)"),
    ("rent_low", "Rent Low"),
    ("rent_high", "Rent High"),
    ("available_date", "Available Date"),
    ("lease_term", "Lease Term"),
    # 2026-05-21 fix: the prior single "Concessions" column collapsed three
    # distinct V2 schema fields into one stringified cell, masking which
    # rows had clean-vs-dirty concession text and which had nothing at
    # all. Split into three explicit columns so reviewers can see the
    # raw producer string, the cleaned variant (may be empty for
    # malformed input), and the quality label. The legacy ``concessions``
    # key is retained on the row dict for backward compatibility with
    # any other consumer, but only the three split columns ship in xlsx.
    ("concession_text", "Concession (Raw)"),
    ("concession_text_clean", "Concession (Cleaned)"),
    ("_concession_quality", "Concession Quality"),
    # 2026-05-22 enrichment: deterministic per-cell signals so reviewers
    # don't have to skim free-form text to figure out what the offer is.
    # Banner is the one-line human render; the other four columns are
    # filter/pivot targets in Sheets.
    ("concession_banner", "Concession (Banner)"),
    ("concession_offer_type", "Offer Type"),
    ("concession_target", "Offer Target"),
    ("concession_value", "Offer Value"),
    ("concession_conditions", "Offer Conditions"),
]

_FAILED_COLUMNS: list[tuple[str, str]] = [
    ("canonical_id", "Canonical ID"),
    ("property_name", "Property Name"),
    ("city", "City"),
    ("state", "State"),
    ("pmc", "Mgmt Company"),
    ("website", "Website"),
    ("ledger_status", "Status"),
    ("carry_forward_used", "Carry-Forward"),
    ("scrape_failed", "Scrape Failed"),
    ("units_count", "Units"),
    ("error_count", "Errors"),
    ("warning_count", "Warnings"),
    ("issue_codes", "Error Codes"),
    ("issue_messages", "Issue Messages (first 3)"),
    ("scrape_outcome", "Scrape Outcome"),
    ("failure_reason", "Failure Reason"),
    ("extraction_tier", "Extract Tier"),
    ("page_load_ms", "Page Load (ms)"),
    ("last_event_at", "Last Scrape Event"),
]


def _write_xlsx(
    out_path: Path,
    sheet_name: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
) -> Path:
    """Write ``rows`` to ``out_path`` with a frozen header row and autofilter.

    openpyxl is pure-Python so this works on Windows ARM64 with no
    native build step. We size columns from a sample of values rather
    than the full column to keep big sheets fast.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover — Workbook always has an active sheet
        ws = wb.create_sheet()
    ws.title = sheet_name[:31]  # Excel caps sheet names at 31 chars

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_align = Alignment(vertical="center")

    headers = [label for _, label in columns]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append([_xlsx_value(row.get(key)) for key, _ in columns])

    ws.freeze_panes = "A2"
    if rows:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    # Approximate column widths from a 200-row sample. Capped at 60 to
    # keep extreme outliers (long failure_reason text) from blowing out
    # the layout — recipients can drag-resize in Sheets / Excel.
    sample = rows[:200]
    for col_idx, (key, label) in enumerate(columns, start=1):
        widest = len(label)
        for row in sample:
            value = row.get(key)
            if value is None:
                continue
            widest = max(widest, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(widest + 2, 60)

    wb.save(out_path)
    log.info("Wrote %s rows to %s", f"{len(rows):,}", out_path)
    return out_path


def _xlsx_value(value: Any) -> Any:
    """Coerce values openpyxl can't natively store into strings."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _truncate(s: str, n: int) -> str:
    s = (s or "").strip().replace("\n", " ").replace("\r", " ")
    if len(s) <= n:
        return s
    return s[: n - 1].rstrip() + "…"


# ── HTML rendering (concise summary; the spreadsheets carry the detail) ─────


def _render_html(
    *,
    run_date: str,
    source_label: str,
    gcs_present: bool,
    scraped_rows: list[dict[str, Any]],
    failed_rows: list[dict[str, Any]],
) -> str:
    """A short HTML body. Big tables go in the XLSX attachments."""
    properties_total = len({r["canonical_id"] for r in scraped_rows if r.get("canonical_id")})
    units_total = sum(1 for r in scraped_rows if r.get("unit_id"))
    failed_total = len(failed_rows)
    success_total = properties_total - failed_total
    success_pct = (100.0 * success_total / properties_total) if properties_total else 0.0

    code_counter: Counter[str] = Counter()
    for fr in failed_rows:
        for code in (fr.get("issue_codes") or "").split(","):
            code = code.strip()
            if code:
                code_counter[code] += 1
    top_codes = code_counter.most_common(10)

    top_failed = sorted(
        failed_rows,
        key=lambda r: (r.get("error_count") or 0, r.get("warning_count") or 0),
        reverse=True,
    )[:10]

    pill_class = "ok" if success_pct >= 95 else "warn" if success_pct >= 80 else "bad"

    css = """
      body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
             color:#1f2937; background:#f4f6fa; margin:0; padding:18px; }
      .wrap { max-width: 760px; margin: 0 auto; background:#fff;
              border:1px solid #e5e7eb; border-radius:10px; overflow:hidden; }
      .header { background: linear-gradient(135deg,#0f172a 0%,#1e3a8a 100%);
                color:#fff; padding:24px 28px; }
      .header h1 { margin:0; font-size:20px; font-weight:700; }
      .header .sub { color:#cbd5e1; font-size:12px; text-transform:uppercase;
                     letter-spacing:0.1em; margin-top:6px; }
      .kpis { padding: 18px 28px; display:flex; flex-wrap:wrap; gap:8px; }
      .kpi { padding:6px 12px; border-radius:999px; font-size:13px;
             background:#1e2632; color:#fff; }
      .kpi.ok { background:#166534; }
      .kpi.warn { background:#92400e; }
      .kpi.bad { background:#991b1b; }
      .source { padding: 0 28px 18px; color:#475569; font-size:12px; }
      .source code { font-family: ui-monospace, Menlo, Consolas, monospace;
                     background:#eff6ff; color:#1e3a8a; padding:1px 6px;
                     border-radius:4px; font-size:11.5px; }
      h2 { margin: 22px 28px 10px; font-size:13px; text-transform:uppercase;
           letter-spacing:0.08em; color:#475569; }
      table { border-collapse: collapse; width: calc(100% - 56px); margin: 0 28px; font-size:13px; }
      th, td { text-align:left; padding:8px 10px; border-bottom:1px solid #f1f5f9; }
      th { background:#f8fafc; font-weight:600; color:#334155; font-size:11px;
           text-transform:uppercase; letter-spacing:0.06em; }
      td.num, th.num { text-align:right; font-variant-numeric: tabular-nums; }
      .empty { margin: 0 28px; padding: 14px 16px; background:#f0fdf4;
               border:1px solid #bbf7d0; border-radius:8px; color:#166534;
               font-weight:500; font-size:13px; }
      code.mono { font-family: ui-monospace, Menlo, Consolas, monospace;
                  background:#eff6ff; color:#1e3a8a; padding:1px 6px;
                  border-radius:4px; font-size:12px; }
      .footer { padding:18px 28px; color:#94a3b8; font-size:11px;
                border-top:1px solid #f1f5f9; background:#fafafa; }
    """

    rows_html = ""
    if top_codes:
        rows_html = "".join(
            f"<tr><td><code class='mono'>{_h(code)}</code></td>"
            f"<td class='num'>{count:,}</td></tr>"
            for code, count in top_codes
        )
        top_codes_table = (
            "<table><thead><tr><th>Issue code</th><th class='num'>Count</th>"
            f"</tr></thead><tbody>{rows_html}</tbody></table>"
        )
    else:
        top_codes_table = "<div class='empty'>No issue codes recorded for this run.</div>"

    if top_failed:
        failed_rows_html = "".join(
            f"<tr>"
            f"<td><code class='mono'>{_h(r.get('canonical_id', ''))}</code></td>"
            f"<td>{_h(r.get('property_name', '') or '—')}</td>"
            f"<td>{_h((r.get('city') or '') + (', ' + r.get('state', '') if r.get('state') else ''))}</td>"
            f"<td>{_h(r.get('ledger_status', ''))}</td>"
            f"<td class='num'>{r.get('error_count') or 0}</td>"
            f"<td>{_h(_truncate(r.get('issue_codes') or '', 60))}</td>"
            f"</tr>"
            for r in top_failed
        )
        top_failed_table = (
            "<table><thead><tr><th>Canonical ID</th><th>Name</th>"
            "<th>Location</th><th>Status</th><th class='num'>Errors</th>"
            f"<th>Codes</th></tr></thead><tbody>{failed_rows_html}</tbody></table>"
        )
    else:
        top_failed_table = "<div class='empty'>✓ All properties succeeded today.</div>"

    gcs_note = (
        "<code>gs://…/runs/{rd}/</code> shards present in Cloud Storage."
        if gcs_present
        else "<code>gs://…/runs/{rd}/</code> not found in Cloud Storage (or BUCKET_NAME unset)."
    ).format(rd=_h(run_date))

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<title>PropAi failures report — {_h(run_date)}</title>
<style>{css}</style>
</head><body><div class="wrap">
  <div class="header">
    <h1>PropAi — Daily failures report</h1>
    <div class="sub">Run {_h(run_date)} · health &amp; debug breadcrumbs</div>
  </div>
  <div class="kpis">
    <span class="kpi">{properties_total:,} properties</span>
    <span class="kpi">{units_total:,} units</span>
    <span class="kpi {pill_class}">{success_pct:.1f}% success</span>
    <span class="kpi {'bad' if failed_total else 'ok'}">{failed_total:,} failed</span>
  </div>
  <div class="source">
    Data source: <code>{_h(source_label)}</code>. {gcs_note}
    <br>Two attachments are included as <code>.xlsx</code> (Gmail
    previews them as Sheets):
    <code>scraped_units_{_h(run_date)}.xlsx</code> with one row per
    captured unit, and
    <code>failed_properties_{_h(run_date)}.xlsx</code> with the failure
    cohort + every error code, scrape outcome, and failure reason.
  </div>
  <h2>Top issue codes</h2>
  {top_codes_table}
  <h2>Top 10 failed properties</h2>
  {top_failed_table}
  <div class="footer">
    Sent automatically by <code class="mono">ma_poc/scripts/email_daily_failures_report.py</code>.
    Reuses the data-provider connection + Gmail MCP send path from
    <code class="mono">email_daily_report.py</code>.
  </div>
</div></body></html>"""


def _h(s: Any) -> str:
    """HTML-escape; tolerates None and non-strings."""
    text = "" if s is None else str(s)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ── Entrypoint ──────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s — %(message)s",
        datefmt="%H:%M:%S",
    )
    load_dotenv(_MA_POC_ROOT / ".env")

    parser = argparse.ArgumentParser(
        description="Email a daily PropAi failures report with two XLSX attachments."
    )
    parser.add_argument("--run-date", default=None,
                        help="YYYY-MM-DD. Defaults to today UTC, then latest run.")
    parser.add_argument("--recipients", default=None,
                        help="Comma-separated. Defaults to REPORT_RECIPIENTS env var.")
    parser.add_argument("--out-dir", default=None,
                        help="Directory for the XLSX files. "
                             "Defaults to ma_poc/data/runs/{run_date}/.")
    parser.add_argument("--use-gcs", action="store_true",
                        help="Force the GCS fallback even when Cloud SQL has rows.")
    parser.add_argument("--local-dir", default=None,
                        help="Read shard artifacts from a local mirror dir "
                             "(e.g. C:/tmp/run-2026-05-21/). Skips Cloud SQL "
                             "and GCS entirely. Per the playbook §Phase 1 — "
                             "use when the run is already downloaded locally.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Write the XLSX files + print HTML; do not send.")
    parser.add_argument("--database-url", default=None,
                        help="Override DATABASE_URL (asyncpg auto-translates to pg8000).")
    args = parser.parse_args(argv)

    bucket = (os.getenv("BUCKET_NAME") or "").strip() or None

    recipients = _parse_recipients(args.recipients) or _parse_recipients(
        os.getenv("REPORT_RECIPIENTS")
    )
    if not recipients and not args.dry_run:
        log.error("No recipients. Set REPORT_RECIPIENTS in .env or pass --recipients.")
        return 1

    sender_name = os.getenv("REPORT_SENDER_NAME") or "PropAi Daily Reports"

    # Local-mirror short-circuit: when --local-dir is given, skip both
    # Cloud SQL and GCS entirely. The playbook (§Phase 1) recommends this
    # mode when the run has already been mirrored to c:/tmp/run-<date>/.
    if args.local_dir:
        local_dir = Path(args.local_dir)
        # Derive run_date from --run-date if given, else infer from the
        # directory name (`run-YYYY-MM-DD` convention). Last resort: today.
        if args.run_date:
            run_date = args.run_date
        else:
            name = local_dir.name
            if name.startswith("run-") and len(name) >= 14:
                run_date = name[4:14]
            else:
                run_date = datetime.now(timezone.utc).date().isoformat()
                log.warning(
                    "Could not derive run_date from %s; defaulting to %s. "
                    "Pass --run-date explicitly to override.",
                    local_dir, run_date,
                )
        log.info("Local mirror mode: dir=%s run_date=%s", local_dir, run_date)
        scraped, failed = _fetch_run_from_local_dir(local_dir)
        source_label = f"Local mirror ({local_dir})"
        gcs_present = False
        return _finish_and_send(
            run_date=run_date,
            source_label=source_label,
            gcs_present=gcs_present,
            scraped=scraped,
            failed=failed,
            recipients=recipients,
            sender_name=sender_name,
            out_dir_arg=args.out_dir,
            dry_run=args.dry_run,
        )

    provider = _open_provider(args.database_url)
    log.info("Data provider: %s · BUCKET_NAME=%s", provider.name, bucket or "(unset)")

    try:
        run_date, gcs_present = _resolve_run_date(provider, args.run_date, bucket=bucket)
        log.info("Resolved run_date=%s · gcs_present=%s", run_date, gcs_present)

        scraped: list[dict[str, Any]] = []
        failed: list[dict[str, Any]] = []
        source_label = "Cloud SQL"

        if not args.use_gcs:
            with Session(provider.engine, future=True) as s:
                # ``select(SomeColumn)`` + ``.scalars()`` yields the column
                # values directly (not row objects) — the result is the cid
                # string, so we collect strings into the set.
                cid_set: set[str] = {
                    cid for cid in s.execute(
                        select(PropertySnapshotRow.canonical_id)
                        .where(PropertySnapshotRow.run_date == run_date)
                    ).scalars()
                    if cid
                }
                # Also include cids from the ledger (failed runs may not have
                # written a snapshot — e.g. fetch HARD_FAIL).
                for cid in s.execute(
                    select(RunLedgerRow.canonical_id)
                    .where(RunLedgerRow.run_date == run_date)
                ).scalars():
                    if cid:
                        cid_set.add(cid)
                meta = _fetch_property_metadata(s, list(cid_set))
                scraped = _fetch_scraped_units_from_sql(s, run_date, meta)
                failed = _fetch_failed_properties_from_sql(s, run_date, meta)

        # Fallback: SQL came up empty (retention sweep) or operator forced GCS.
        if (args.use_gcs or not scraped) and bucket and gcs_present:
            log.warning(
                "%s — falling back to GCS at gs://%s/runs/%s/",
                "User requested --use-gcs" if args.use_gcs else "SQL returned no rows",
                bucket, run_date,
            )
            scraped, failed = _fetch_run_from_gcs(bucket, run_date)
            source_label = f"Cloud Storage (gs://{bucket}/runs/{run_date}/)"
        elif not scraped and not failed:
            log.warning(
                "No data for run_date=%s in Cloud SQL or GCS. "
                "Email will be empty but will still send.", run_date,
            )

    finally:
        provider.close()

    return _finish_and_send(
        run_date=run_date,
        source_label=source_label,
        gcs_present=gcs_present,
        scraped=scraped,
        failed=failed,
        recipients=recipients,
        sender_name=sender_name,
        out_dir_arg=args.out_dir,
        dry_run=args.dry_run,
    )


def _finish_and_send(
    *,
    run_date: str,
    source_label: str,
    gcs_present: bool,
    scraped: list[dict[str, Any]],
    failed: list[dict[str, Any]],
    recipients: list[str] | None,
    sender_name: str,
    out_dir_arg: str | None,
    dry_run: bool,
) -> int:
    """Write the two XLSX attachments and send the email.

    Shared tail between the Cloud SQL / GCS path and the --local-dir path
    so the rendered HTML, file naming, and send-or-dry-run semantics
    stay identical no matter how the rows were sourced.
    """
    if out_dir_arg:
        out_dir = Path(out_dir_arg)
    else:
        out_dir = _MA_POC_ROOT / "data" / "runs" / run_date
    out_dir.mkdir(parents=True, exist_ok=True)

    scraped_path = out_dir / f"scraped_units_{run_date}.xlsx"
    failed_path = out_dir / f"failed_properties_{run_date}.xlsx"
    _write_xlsx(scraped_path, "Scraped units", _SCRAPED_COLUMNS, scraped)
    _write_xlsx(failed_path, "Failed properties", _FAILED_COLUMNS, failed)

    log.info(
        "run_date=%s scraped_units=%s failed_props=%s",
        run_date, f"{len(scraped):,}", f"{len(failed):,}",
    )

    html = _render_html(
        run_date=run_date,
        source_label=source_label,
        gcs_present=gcs_present,
        scraped_rows=scraped,
        failed_rows=failed,
    )
    subject = f"PropAi failures report — {run_date} ({len(failed)} failed)"

    try:
        result = send_html_email(
            subject=subject,
            html_body=html,
            recipients=recipients or None,
            sender_name=sender_name,
            attachments=[scraped_path, failed_path],
            dry_run=dry_run,
        )
    except Exception as exc:
        log.exception("Send failed: %s", exc)
        return 1

    if result.get("isError"):
        log.error("Gmail MCP returned an error: %s", result.get("content"))
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
